"""
ESPANSIONE SEQUENZIALE 5 PUNTI VENDITA CON BUDGET LIMITATO
============================================================

Simula l'apertura sequenziale di 5 punti vendita partendo da un budget
iniziale di €250.000, calcolando quando è possibile aprire ogni nuovo
negozio basandosi sul cash flow cumulativo generato.

LOGICA:
- Inizio con €250.000 di capitale
- Apro PdV1 subito (M1) con investimento di €170.000
- Rimangono €80.000 di liquidità
- Ogni mese, accumulo CF dai negozi aperti
- Quando CF cumulativo ≥ €170.000, posso aprire il prossimo negozio

PARAMETRI:
- Investimento per PdV: €170.000
- EBITDA mensile a regime: €51.530
- Ricavi mensili a regime: €107.550
- COGS %: 39,7%
- OPEX fissi per PdV: €13.345
- Ramp-up: 30% (M1-2), 60% (M3-9), 100% (M10+)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crea_simulazione_espansione():
    """
    Crea un file Excel che simula l'espansione da 1 a 5 PdV con budget di 250k
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Espansione Sequenziale"

    # Stili
    title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="3D7AB8", end_color="3D7AB8", fill_type="solid")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3D7AB8", end_color="3D7AB8", fill_type="solid")
    subheader_font = Font(name='Arial', size=11, bold=True, color="000000")
    subheader_fill = PatternFill(start_color="FFF4E0", end_color="FFF4E0", fill_type="solid")
    total_font = Font(name='Arial', size=11, bold=True)
    param_font = Font(name='Arial', size=10, italic=True, color="595959")
    highlight_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def apply_style(cell, font, fill=None, alignment=None):
        cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment

    # ==================================================================================
    # PARAMETRI
    # ==================================================================================

    budget_iniziale = 250000
    investimento_pdv = 170000
    ebitda_regime = 51530
    ricavi_regime = 107550
    cogs_perc = 0.397
    opex_fissi = 13345

    def calcola_rampup(mese_da_apertura):
        """Calcola il coefficiente di ramp-up in base ai mesi dall'apertura"""
        if mese_da_apertura <= 2:
            return 0.30
        elif mese_da_apertura <= 9:
            return 0.60
        else:
            return 1.00

    # ==================================================================================
    # SEZIONE 1: TITOLO E PARAMETRI
    # ==================================================================================

    ws['A1'] = "SIMULAZIONE ESPANSIONE SEQUENZIALE - 5 PUNTI VENDITA CON BUDGET €250.000"
    ws.merge_cells('A1:L1')
    apply_style(ws['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35

    row = 3
    ws.cell(row, 1, "PARAMETRI SIMULAZIONE").font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    apply_style(ws.cell(row, 1), header_font, header_fill, Alignment(horizontal='center'))
    row += 1

    params_input = [
        ["Parametro", "Valore", "Note"],
        ["Budget Iniziale", budget_iniziale, "Capitale disponibile al mese 1"],
        ["Investimento per PdV", investimento_pdv, "Costo apertura singolo negozio"],
        ["Liquidità dopo PdV1", budget_iniziale - investimento_pdv, "Capitale residuo dopo primo negozio"],
        [],
        ["PARAMETRI ECONOMICI PER PdV"],
        ["Ricavi Mensili a Regime", ricavi_regime, "100% capacità"],
        ["COGS %", f"{cogs_perc:.1%}", "Costi variabili su ricavi"],
        ["OPEX Fissi Mensili", opex_fissi, "Costi fissi per negozio"],
        ["EBITDA Mensile a Regime", ebitda_regime, "Margine operativo a regime"],
        [],
        ["RAMP-UP"],
        ["Mesi 1-2", "30%", "Fase iniziale"],
        ["Mesi 3-9", "60%", "Fase crescita"],
        ["Mesi 10+", "100%", "Regime"],
    ]

    for r_idx, row_data in enumerate(params_input, row):
        if not row_data:
            continue
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(r_idx, c_idx, val)
            if r_idx == row:
                apply_style(cell, subheader_font, subheader_fill)
            elif len(row_data) > 1 and c_idx == 1:
                if "PARAMETRI" in str(val) or "RAMP-UP" in str(val):
                    cell.font = Font(bold=True)
        if len(row_data) >= 2 and isinstance(row_data[1], (int, float)):
            ws.cell(r_idx, 2).number_format = '€ #,##0'

    row = row + len([r for r in params_input if r]) + 2

    # ==================================================================================
    # SEZIONE 2: SIMULAZIONE MENSILE
    # ==================================================================================

    ws.cell(row, 1, "PROIEZIONE MENSILE - ESPANSIONE SEQUENZIALE").font = header_font
    ws.merge_cells(f'A{row}:L{row}')
    apply_style(ws.cell(row, 1), header_font, header_fill, Alignment(horizontal='center'))
    row += 1

    # Simula mese per mese
    mesi_simulazione = 36  # 3 anni

    # Stato della simulazione
    pdv_aperti = []  # Lista di tuple (numero_pdv, mese_apertura)
    cf_cumulativo = budget_iniziale

    # Dati mensili
    dati_mensili = []

    for mese in range(1, mesi_simulazione + 1):
        # Verifica se possiamo aprire un nuovo PdV
        if mese == 1:
            # Apri PdV1 al mese 1
            pdv_aperti.append((1, 1))
            cf_cumulativo -= investimento_pdv
        elif len(pdv_aperti) < 5 and cf_cumulativo >= investimento_pdv:
            # Possiamo aprire un nuovo PdV
            nuovo_pdv = len(pdv_aperti) + 1
            pdv_aperti.append((nuovo_pdv, mese))
            cf_cumulativo -= investimento_pdv

        # Calcola CF del mese corrente
        ricavi_mese = 0
        cogs_mese = 0
        opex_mese = 0

        for num_pdv, mese_apertura in pdv_aperti:
            mesi_da_apertura = mese - mese_apertura + 1
            rampup = calcola_rampup(mesi_da_apertura)

            ricavi_pdv = ricavi_regime * rampup
            cogs_pdv = ricavi_pdv * cogs_perc
            opex_pdv = opex_fissi

            ricavi_mese += ricavi_pdv
            cogs_mese += cogs_pdv
            opex_mese += opex_pdv

        cf_mese = ricavi_mese - cogs_mese - opex_mese
        cf_cumulativo += cf_mese

        # Evento del mese (apertura nuovo PdV?)
        evento = ""
        for num_pdv, mese_apertura in pdv_aperti:
            if mese_apertura == mese:
                evento = f"🔓 APERTURA PdV{num_pdv}"

        dati_mensili.append({
            'mese': mese,
            'pdv_attivi': len(pdv_aperti),
            'ricavi': ricavi_mese,
            'cogs': cogs_mese,
            'opex': opex_mese,
            'cf_mese': cf_mese,
            'cf_cumul': cf_cumulativo,
            'evento': evento
        })

    # Scrivi header tabella
    headers_tabella = ["Mese", "PdV Attivi", "Ricavi Tot", "COGS Tot", "OPEX Tot",
                      "CF Mensile", "CF Cumulativo", "Evento"]
    for c_idx, h in enumerate(headers_tabella, 1):
        cell = ws.cell(row, c_idx, h)
        apply_style(cell, subheader_font, subheader_fill)
    row_header = row
    row += 1

    # Scrivi dati mensili
    for dati in dati_mensili:
        ws.cell(row, 1, dati['mese'])
        ws.cell(row, 2, dati['pdv_attivi'])
        ws.cell(row, 3, dati['ricavi'])
        ws.cell(row, 4, dati['cogs'])
        ws.cell(row, 5, dati['opex'])
        ws.cell(row, 6, dati['cf_mese'])
        ws.cell(row, 7, dati['cf_cumul'])
        ws.cell(row, 8, dati['evento'])

        # Formattazione
        for c in [3, 4, 5, 6, 7]:
            ws.cell(row, c).number_format = '€ #,##0'

        # Colora CF cumulativo
        if dati['cf_cumul'] >= 0:
            ws.cell(row, 7).fill = green_fill
        else:
            ws.cell(row, 7).fill = red_fill

        # Evidenzia righe di apertura
        if dati['evento']:
            ws.cell(row, 8).fill = highlight_fill
            ws.cell(row, 8).font = total_font

        row += 1

    row += 1

    # ==================================================================================
    # SEZIONE 3: TIMELINE APERTURE
    # ==================================================================================

    ws.cell(row, 1, "TIMELINE APERTURE E RISULTATI").font = header_font
    ws.merge_cells(f'A{row}:F{row}')
    apply_style(ws.cell(row, 1), header_font, header_fill, Alignment(horizontal='center'))
    row += 1

    # Riepilogo aperture
    headers_aperture = ["Punto Vendita", "Mese Apertura", "CF Cumul Prima", "CF Cumul Dopo", "Tempo da Precedente"]
    for c_idx, h in enumerate(headers_aperture, 1):
        cell = ws.cell(row, c_idx, h)
        apply_style(cell, subheader_font, subheader_fill)
    row += 1

    mese_prev = 0
    for num_pdv, mese_apertura in pdv_aperti:
        # Trova CF cumulativo prima e dopo apertura
        cf_prima = None
        cf_dopo = None

        for dati in dati_mensili:
            if dati['mese'] == mese_apertura:
                # CF dopo apertura (già include l'investimento)
                cf_dopo = dati['cf_cumul']
                # CF prima = CF dopo + investimento
                cf_prima = cf_dopo + investimento_pdv
                break

        tempo_da_prev = mese_apertura - mese_prev if mese_prev > 0 else 0

        ws.cell(row, 1, f"PdV {num_pdv}")
        ws.cell(row, 2, f"Mese {mese_apertura}")
        ws.cell(row, 3, cf_prima if cf_prima else "")
        ws.cell(row, 4, cf_dopo if cf_dopo else "")
        ws.cell(row, 5, f"{tempo_da_prev} mesi" if tempo_da_prev > 0 else "Inizio")

        for c in [3, 4]:
            if ws.cell(row, c).value:
                ws.cell(row, c).number_format = '€ #,##0'

        ws.cell(row, 1).font = total_font

        mese_prev = mese_apertura
        row += 1

    row += 1

    # ==================================================================================
    # SEZIONE 4: CONCLUSIONI
    # ==================================================================================

    ws.cell(row, 1, "ANALISI E CONCLUSIONI").font = header_font
    ws.merge_cells(f'A{row}:F{row}')
    apply_style(ws.cell(row, 1), header_font, header_fill, Alignment(horizontal='center'))
    row += 1

    # Calcola metriche finali
    cf_finale = dati_mensili[-1]['cf_cumul']
    mesi_totali = mesi_simulazione
    num_pdv_finale = len(pdv_aperti)
    ultimo_mese_apertura = pdv_aperti[-1][1] if pdv_aperti else 0

    conclusioni = [
        ["RISULTATI FINALI"],
        [f"Budget iniziale:", f"€{budget_iniziale:,}"],
        [f"Numero PdV aperti:", f"{num_pdv_finale} su 5"],
        [f"Ultimo PdV aperto al:", f"Mese {ultimo_mese_apertura}"],
        [f"Cash Flow finale (M{mesi_totali}):", f"€{cf_finale:,.0f}"],
        [f"Capitale investito totale:", f"€{num_pdv_finale * investimento_pdv:,}"],
        [f"Capitale generato:", f"€{cf_finale - budget_iniziale:,.0f}"],
        [],
        ["INTERPRETAZIONE"],
        ["Con un budget iniziale di €250.000:"],
        [f"- Puoi aprire immediatamente il primo negozio (€170k)"],
        [f"- Rimangono €80k di liquidità operativa"],
        [f"- Ogni nuovo negozio si apre quando CF cumul ≥ €170k"],
        [f"- L'autofinanziamento permette espansione senza capitale aggiuntivo"],
        [f"- Tempo medio tra aperture: ~{ultimo_mese_apertura//(num_pdv_finale-1) if num_pdv_finale > 1 else 0} mesi"],
    ]

    for conclusione in conclusioni:
        if not conclusione:
            row += 1
            continue

        if len(conclusione) == 1:
            ws.cell(row, 1, conclusione[0])
            ws.cell(row, 1).font = Font(bold=True, size=11)
            if "RISULTATI" in conclusione[0]:
                ws.cell(row, 1).fill = highlight_fill
            row += 1
        else:
            ws.cell(row, 1, conclusione[0])
            ws.cell(row, 2, conclusione[1])
            if "Budget" in conclusione[0] or "finale" in conclusione[0]:
                ws.cell(row, 1).font = total_font
                ws.cell(row, 2).font = total_font
            row += 1

    # Larghezze colonne
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30

    # Salva file
    filename = "Espansione_Sequenziale_5PdV.xlsx"
    wb.save(filename)
    wb.close()

    # Prepara output
    aperture_info = [(num, mese) for num, mese in pdv_aperti]

    return filename, aperture_info, cf_finale

if __name__ == "__main__":
    filename, aperture, cf_finale = crea_simulazione_espansione()

    print("\n" + "="*80)
    print("✅ SIMULAZIONE ESPANSIONE SEQUENZIALE COMPLETATA")
    print("="*80)
    print(f"\n📊 File generato: {filename}")
    print(f"\n🎯 RISULTATI:")
    print(f"   Budget iniziale: €250.000")
    print(f"   Costo per PdV: €170.000")
    print(f"\n📅 TIMELINE APERTURE:")

    for num_pdv, mese_apertura in aperture:
        print(f"   • PdV {num_pdv}: Mese {mese_apertura}")

    print(f"\n💰 CASH FLOW FINALE (Mese 36): €{cf_finale:,.0f}")
    print(f"\n💡 INTERPRETAZIONE:")
    print(f"   Con €250k puoi aprire {len(aperture)} negozi in {aperture[-1][1] if aperture else 0} mesi")
    print(f"   L'espansione è completamente autofinanziata dal cash flow generato!")
    print("\n" + "="*80)
