"""
OTTIMIZZAZIONE TIMING APERTURE - Italian Corner
================================================

Questo script analizza quando conviene aprire il secondo punto vendita
per minimizzare l'investimento iniziale richiesto, tenendo conto di:
- Ricavi mensili con ramp-up progressivo
- Costi variabili (COGS) proporzionali ai ricavi
- Costi fissi mensili (OPEX)
- Cash flow cumulativo

LOGICA:
Il secondo negozio dovrebbe aprire quando il cash flow cumulativo del primo
è sufficientemente positivo da coprire (parzialmente o totalmente) l'investimento
del secondo, riducendo così il fabbisogno di capitale iniziale.

FORMULA OTTIMALE:
Mese ottimale apertura PdV2 = MIN(Investimento Totale Richiesto)
dove Investimento Totale = MAX(valore assoluto del minimo CF cumulativo)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crea_analisi_timing_aperture():
    """
    Crea un file Excel che analizza diversi scenari di timing per l'apertura
    del secondo punto vendita (da mese 6 a mese 18)
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Analisi Timing Aperture"

    # Stili
    title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="3D7AB8", end_color="3D7AB8", fill_type="solid")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3D7AB8", end_color="3D7AB8", fill_type="solid")
    subheader_font = Font(name='Arial', size=11, bold=True, color="000000")
    subheader_fill = PatternFill(start_color="FFF4E0", end_color="FFF4E0", fill_type="solid")
    total_font = Font(name='Arial', size=11, bold=True)
    param_font = Font(name='Arial', size=10, italic=True, color="595959")
    highlight_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    def apply_style(cell, font, fill=None, alignment=None):
        cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment

    # ==================================================================================
    # SEZIONE 1: PARAMETRI DI INPUT
    # ==================================================================================

    ws['A1'] = "ANALISI OTTIMIZZAZIONE TIMING APERTURA SECONDO PUNTO VENDITA"
    ws.merge_cells('A1:H1')
    apply_style(ws['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30

    row = 3
    ws.cell(row, 1, "PARAMETRI DI INPUT").font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    apply_style(ws.cell(row, 1), header_font, header_fill, Alignment(horizontal='center'))
    row += 1

    # Parametri economici
    params_input = [
        ["Parametro", "Valore", "Unità", "Note"],
        ["Investimento per PdV", "=RIFERIMENTO_ESTERNO", "€", "Da Business_Plan_Italian_Corner.xlsx"],
        ["EBITDA Mensile a Regime", 51530, "€", "Margine operativo mensile al 100% capacità"],
        ["Ricavi Mensili a Regime", 107550, "€", "Fatturato al 100% capacità"],
        ["COGS %", 0.397, "%", "Costi variabili su ricavi"],
        ["OPEX Fissi Mensili", 13345, "€", "Costi operativi fissi per PdV"],
        [],
        ["COEFFICIENTI RAMP-UP"],
        ["Mesi 1-2", 0.30, "%", "30% capacità - fase iniziale"],
        ["Mesi 3-9", 0.60, "%", "60% capacità - fase crescita"],
        ["Mesi 10+", 1.00, "%", "100% capacità - regime"],
    ]

    for r_idx, row_data in enumerate(params_input, row):
        if not row_data:
            continue
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(r_idx, c_idx, val)
            if r_idx == row:
                apply_style(cell, subheader_font, subheader_fill)
            elif "COEFFICIENTI" in str(val):
                cell.font = Font(bold=True)
        if len(row_data) >= 2 and isinstance(row_data[1], (int, float)) and row_data[1] != 0:
            ws.cell(r_idx, 2).number_format = '€ #,##0' if '€' in str(row_data[2]) else '0.0%'

    row = row + len([r for r in params_input if r]) + 2

    # ==================================================================================
    # SEZIONE 2: ANALISI SCENARI DI TIMING
    # ==================================================================================

    ws.cell(row, 1, "ANALISI SCENARI: APERTURA SECONDO PUNTO VENDITA").font = header_font
    ws.merge_cells(f'A{row}:P{row}')
    apply_style(ws.cell(row, 1), header_font, header_fill, Alignment(horizontal='center'))
    row += 1

    # Headers
    headers = ["Mese Apertura PdV2", "CF Cumul PdV1", "Invest PdV2", "Gap da Coprire",
               "Capitale Tot Inizio", "CF Min Post", "Capitale Aggiuntivo", "Risparmio vs Simult", "% Risparmio", "Raccomandazione"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row, c_idx, h)
        apply_style(cell, subheader_font, subheader_fill)
    row_headers = row
    row += 1

    # Parametri di riferimento
    investimento = 166695  # Valore tipico
    ebitda_regime = 51530
    ricavi_regime = 107550
    cogs_perc = 0.397
    opex_fissi = 13345

    def calcola_rampup(mese_da_apertura):
        """Calcola il coefficiente di ramp-up in base ai mesi dall'apertura"""
        if mese_da_apertura <= 2:
            return 0.30
        elif mese_da_apertura <= 9:
            return 0.60
        else:
            return 1.00

    def calcola_cf_pdv1(fino_a_mese):
        """Calcola il cash flow cumulativo del PdV1 dal mese 1 fino al mese specificato"""
        cf_cumul = -investimento  # Investimento iniziale al mese 1

        for mese in range(1, fino_a_mese + 1):
            rampup = calcola_rampup(mese)
            ricavi_mese = ricavi_regime * rampup
            cogs_mese = ricavi_mese * cogs_perc
            cf_mese = ricavi_mese - cogs_mese - opex_fissi
            cf_cumul += cf_mese

        return cf_cumul

    # Analizza diversi mesi di apertura (da mese 6 a mese 18)
    row_start_scenari = row
    migliore_mese = None
    minimo_fabbisogno = float('inf')

    scenari_data = []

    for mese_apertura_pdv2 in range(6, 19):  # Da mese 6 a mese 18
        # CF cumulativo PdV1 al momento dell'apertura PdV2 (appena prima)
        cf_pdv1_pre = calcola_cf_pdv1(mese_apertura_pdv2 - 1)

        # Investimento PdV2
        invest_pdv2 = investimento

        # Gap da coprire: quanto manca per pagare il PdV2?
        # Se CF_PdV1 è positivo e > investimento, gap = 0
        # Se CF_PdV1 è negativo o < investimento, gap = investimento - CF_PdV1
        gap_da_coprire = max(0, invest_pdv2 - max(0, cf_pdv1_pre))

        # Capitale totale all'inizio (mese 1):
        # = Investimento PdV1 + Gap da coprire per PdV2
        capitale_tot_inizio = investimento + gap_da_coprire

        # CF dopo investimento PdV2
        cf_post_invest = cf_pdv1_pre - invest_pdv2

        # Simula i mesi successivi fino a 24 mesi per trovare il minimo CF cumulativo
        cf_cumul = cf_post_invest
        cf_min = cf_cumul
        mese_min = mese_apertura_pdv2

        for mese in range(mese_apertura_pdv2, 25):
            # CF PdV1
            rampup_pdv1 = calcola_rampup(mese)
            ricavi_pdv1 = ricavi_regime * rampup_pdv1
            cogs_pdv1 = ricavi_pdv1 * cogs_perc
            cf_pdv1 = ricavi_pdv1 - cogs_pdv1 - opex_fissi

            # CF PdV2
            mesi_da_apertura_pdv2 = mese - mese_apertura_pdv2 + 1
            rampup_pdv2 = calcola_rampup(mesi_da_apertura_pdv2)
            ricavi_pdv2 = ricavi_regime * rampup_pdv2
            cogs_pdv2 = ricavi_pdv2 * cogs_perc
            cf_pdv2 = ricavi_pdv2 - cogs_pdv2 - opex_fissi

            cf_cumul += cf_pdv1 + cf_pdv2

            if cf_cumul < cf_min:
                cf_min = cf_cumul
                mese_min = mese

        # Capitale aggiuntivo necessario oltre ai 170k già investiti
        # Se il CF min post-apertura è negativo e peggiore del gap già coperto,
        # serve altro capitale
        capitale_aggiuntivo = max(0, abs(cf_min) - capitale_tot_inizio)
        capitale_tot_reale = capitale_tot_inizio + capitale_aggiuntivo

        # Risparmio vs apertura simultanea (2 x 170k = 340k)
        risparmio = (2 * investimento) - capitale_tot_reale
        perc_risparmio = (risparmio / (2 * investimento)) * 100 if capitale_tot_reale > 0 else 100

        # Raccomandazione (calcolata dopo aver analizzato tutti gli scenari)
        raccomandazione = ""  # Verrà impostata dopo

        scenari_data.append({
            'mese': mese_apertura_pdv2,
            'cf_pdv1': cf_pdv1_pre,
            'invest_pdv2': invest_pdv2,
            'gap': gap_da_coprire,
            'capitale_inizio': capitale_tot_inizio,
            'cf_min': cf_min,
            'capitale_agg': capitale_aggiuntivo,
            'capitale_tot': capitale_tot_reale,
            'risparmio': risparmio,
            'perc_risp': perc_risparmio,
            'racc': raccomandazione
        })

        # Aggiorna il minimo (capitale totale reale)
        if capitale_tot_reale < minimo_fabbisogno:
            minimo_fabbisogno = capitale_tot_reale
            migliore_mese = mese_apertura_pdv2

    # Imposta la raccomandazione per lo scenario ottimale
    for scenario in scenari_data:
        if scenario['mese'] == migliore_mese:
            scenario['racc'] = "★ OTTIMALE ★"

    # Scrivi i dati degli scenari
    for scenario in scenari_data:
        ws.cell(row, 1, scenario['mese'])
        ws.cell(row, 2, scenario['cf_pdv1'])
        ws.cell(row, 3, scenario['invest_pdv2'])
        ws.cell(row, 4, scenario['gap'])
        ws.cell(row, 5, scenario['capitale_inizio'])
        ws.cell(row, 6, scenario['cf_min'])
        ws.cell(row, 7, scenario['capitale_agg'])
        ws.cell(row, 8, scenario['risparmio'])
        ws.cell(row, 9, scenario['perc_risp'])
        ws.cell(row, 10, scenario['racc'])

        # Formattazione
        for c in [2, 3, 4, 5, 6, 7, 8]:
            ws.cell(row, c).number_format = '€ #,##0'
        ws.cell(row, 9).number_format = '0.0%'

        # Evidenzia riga ottimale
        if scenario['racc']:
            for c in range(1, 11):
                ws.cell(row, c).fill = highlight_fill
                ws.cell(row, c).font = total_font

        row += 1

    row += 1

    # ==================================================================================
    # SEZIONE 3: CONCLUSIONI E SPIEGAZIONI
    # ==================================================================================

    ws.cell(row, 1, "CONCLUSIONI E RACCOMANDAZIONI").font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    apply_style(ws.cell(row, 1), header_font, header_fill, Alignment(horizontal='center'))
    row += 1

    conclusioni = [
        ["RISULTATO OTTIMALE"],
        [f"Mese ottimale apertura PdV2:", f"Mese {migliore_mese}"],
        [f"Capitale totale richiesto (M1):", f"€{minimo_fabbisogno:,.0f}"],
        [f"Risparmio vs apertura simultanea:", f"€{(2*investimento - minimo_fabbisogno):,.0f} ({((2*investimento - minimo_fabbisogno)/(2*investimento)*100):.1f}%)"],
        [],
        ["SPIEGAZIONE LOGICA"],
        ["Il timing ottimale bilancia due fattori:"],
        ["1. CASH FLOW ACCUMULATO: Aspettare permette al PdV1 di generare più cassa"],
        ["2. RAMP-UP: Il PdV2 impiega 10 mesi per raggiungere il 100% capacità"],
        ["3. EFFETTO COMBINATO: Il CF positivo del PdV1 deve coprire il periodo di ramp-up del PdV2"],
        [],
        ["Se apri TROPPO PRESTO:"],
        ["- Il PdV1 non ha ancora generato abbastanza cassa positiva"],
        ["- Devi investire quasi il capitale completo per entrambi"],
        [],
        ["Se apri TROPPO TARDI:"],
        ["- Perdi opportunità di ricavi del PdV2"],
        ["- L'effetto moltiplicativo dei due negozi si ritarda"],
        [],
        ["FORMULA CORRETTA UTILIZZATA"],
        ["Per ogni mese potenziale di apertura:"],
        ["1. CF_PdV1(M) = Cash flow cumulativo PdV1 fino al mese M-1"],
        ["2. Gap = MAX(0, Invest_PdV2 - CF_PdV1) = quanto manca per pagare PdV2"],
        ["3. Capitale_Inizio = Invest_PdV1 + Gap = capitale da avere al mese 1"],
        ["4. Simula CF combinato da M a M24, trova il minimo"],
        ["5. Se CF_min < -Capitale_Inizio, serve capitale aggiuntivo"],
        ["6. Capitale_Totale = Capitale_Inizio + eventuali aggiunte"],
        ["7. Il mese con minimo Capitale_Totale è quello ottimale"],
        [],
        ["IMPORTANTE: Capitale minimo assoluto = €170.000 (1 PdV al mese 1)"],
    ]

    for conclusione in conclusioni:
        if not conclusione:
            row += 1
            continue

        if len(conclusione) == 1:
            ws.cell(row, 1, conclusione[0])
            ws.cell(row, 1).font = Font(bold=True, size=11)
            if "RISULTATO" in conclusione[0]:
                ws.cell(row, 1).fill = highlight_fill
            row += 1
        else:
            ws.cell(row, 1, conclusione[0])
            ws.cell(row, 2, conclusione[1])
            if "Mese ottimale" in conclusione[0] or "Investimento reale" in conclusione[0]:
                ws.cell(row, 1).font = total_font
                ws.cell(row, 2).font = total_font
            row += 1

    # Larghezze colonne
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20

    # Salva file
    filename = "Ottimizzazione_Timing_Aperture.xlsx"
    wb.save(filename)
    wb.close()

    return filename, migliore_mese, minimo_fabbisogno

if __name__ == "__main__":
    filename, mese_ottimale, investimento_min = crea_analisi_timing_aperture()

    print("\n" + "="*70)
    print("✅ ANALISI OTTIMIZZAZIONE TIMING COMPLETATA")
    print("="*70)
    print(f"\n📊 File generato: {filename}")
    print(f"\n🎯 RISULTATO OTTIMALE:")
    print(f"   • Mese apertura PdV2: Mese {mese_ottimale}")
    print(f"   • Investimento reale richiesto: €{investimento_min:,.0f}")
    print(f"   • Risparmio: €{(2*166695 - investimento_min):,.0f}")
    print(f"\n💡 INTERPRETAZIONE:")
    print(f"   Aprendo il secondo negozio al mese {mese_ottimale}, il cash flow")
    print(f"   generato dal primo negozio riduce significativamente il capitale")
    print(f"   necessario all'inizio, ottimizzando l'efficienza dell'investimento.")
    print("\n" + "="*70)
