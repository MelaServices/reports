#!/usr/bin/env python3
"""
Business Plan Generator - Italian Corner
Versione avanzata con formule Excel per un modello completamente dinamico.
Le modifiche nel foglio 'Parametri' aggiornano automaticamente l'intero business plan.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crea_business_plan_parametrizzato():
    """
    Crea un file Excel per il business plan di "Italian Corner".
    Il file è strutturato in fogli interconnessi tramite formule.
    """
    wb = Workbook()

    # --- STILI COMUNI (Nuovo Schema Colori) ---
    # Header tabelle: #7A9B5C (Verde oliva) con testo bianco
    # Totali/KPI: #B83A38 (Rosso) con testo bianco
    # Dati positivi: #5FB34A (Verde check)
    # Background alternato: #F5F5F5 (Grigio chiaro)
    # Note/Info: #E8DCC8 (Beige) background, #4A4138 (Marrone) testo

    title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="7A9B5C", end_color="7A9B5C", fill_type="solid")  # Verde oliva
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7A9B5C", end_color="7A9B5C", fill_type="solid")  # Verde oliva
    subheader_font = Font(name='Arial', size=11, bold=True, color="000000")
    subheader_fill = PatternFill(start_color="E8DCC8", end_color="E8DCC8", fill_type="solid")  # Beige

    # Totali e KPI importanti
    total_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="B83A38", end_color="B83A38", fill_type="solid")  # Rosso

    # Dati positivi (revenue, cash flow positivo)
    positive_fill = PatternFill(start_color="5FB34A", end_color="5FB34A", fill_type="solid")  # Verde check
    positive_font = Font(name='Arial', size=10, color="FFFFFF")

    # Dati negativi/attenzione (costi, cash flow negativo)
    negative_fill = PatternFill(start_color="B83A38", end_color="B83A38", fill_type="solid")  # Rosso
    negative_font = Font(name='Arial', size=10, color="FFFFFF")

    # Note e informazioni
    note_fill = PatternFill(start_color="E8DCC8", end_color="E8DCC8", fill_type="solid")  # Beige
    note_font = Font(name='Arial', size=10, italic=True, color="4A4138")  # Marrone
    param_font = Font(name='Arial', size=10, italic=True, color="4A4138")  # Marrone
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Helper function per applicare stili
    def apply_style(cell, font, fill=None, alignment=None, border=None, number_format=None):
        cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format

    # Logo rimosso come da richiesta utente

    # ==============================================================================
    # 1. FOGLIO PARAMETRI (IL "CRUSCOTTO")
    # ==============================================================================
    ws_params = wb.active
    ws_params.title = "Parametri"
    
    ws_params['A1'] = "CRUSCOTTO PARAMETRI DEL MODELLO"
    ws_params.merge_cells('A1:F1')
    apply_style(ws_params['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws_params.row_dimensions[1].height = 30

    # Aggiungi logo

    params_data = [
        ["VOLUMI DI VENDITA (MEDIA GIORNALIERA)", None, None, "SCENARI"],
        ["Descrizione", "Valore Base", "Unità", "Pessimistico", "Ottimistico"],
        ["Piatti venduti in store", 250, "pz/giorno", 200, 400],
        ["Piatti venduti in delivery", 120, "pz/giorno", 80, 180],
        ["Bevande (bibite)", 150, "pz/giorno", 100, 220],
        ["Bevande (acqua)", 250, "pz/giorno", 180, 320],
        [],
        ["PREZZI DI VENDITA E COSTI VARIABILI (PER UNITÀ)"],
        ["Descrizione", "Prezzo IVA Incl. (€)", "Prezzo IVA Escl. (€)", "Costo Unità (€)", "% Margine"],
        ["Piatto pronto", "=C11*1.10", 7.27, 3.50, "=IFERROR((C11-D11)/C11,0)"],
        ["Bevanda (bibita)", "=C12*1.10", 2.27, 0.60, "=IFERROR((C12-D12)/C12,0)"],
        ["Bevanda (acqua)", "=C13*1.10", 0.91, 0.15, "=IFERROR((C13-D13)/C13,0)"],
        [],
        ["COSTI OPERATIVI (OPEX MENSILI)"],
        ["Descrizione", "Importo Mensile (€)", "Note"],
        ["Affitto locale", 3000, "Locale 25mq in zona strategica"],
        ["Personale (costo azienda)", "=B20*B19", "Calcolato: addetti x costo per addetto"],
        ["Numero addetti", 2, "Turni per coprire 10-12 ore"],
        ["Costo lordo per addetto", 3400, "Include stipendio, tasse, contributi"],
        ["Trasporti e logistica", 800, "Consegna giornaliera materie prime"],
        ["Utenze (energia, acqua, web)", 700, "Consumo intensivo microonde"],
        ["Marketing e pubblicità", 1500, "Campagne social, local marketing"],
        ["Assicurazioni e varie", 500, "RC, commercialista, etc."],
        [],
        ["PARAMETRI GENERALI"],
        ["Descrizione", "Valore", "Unità"],
        ["Giorni lavorativi al mese", 30, "giorni"],
        ["Commissione delivery", 0.28, "% sul prezzo"],  # Aggiornato da 0.20 a 0.28
    ]

    for r_idx, row_data in enumerate(params_data, 1):
        if not row_data:
            # Riga vuota - skip ma mantieni l'indice allineato
            continue
        is_header_row = r_idx in [1, 8, 14, 25]  # Aggiornato per riflettere la struttura corretta
        for c_idx, cell_data in enumerate(row_data, 1):
            # Evita di scrivere nelle celle che verranno unite: solo la cella A della riga di intestazione deve avere un valore
            if is_header_row and c_idx > 1:
                continue
            cell = ws_params.cell(row=r_idx + 1, column=c_idx, value=cell_data)
            if isinstance(cell_data, str) and cell_data.startswith("="):
                cell.value = cell_data
            if r_idx in [2, 9, 15, 26]:  # Righe subheader aggiornate
                apply_style(cell, subheader_font, subheader_fill)
        # Esegui il merge e applica lo stile dopo aver inserito il valore nella cella A
        if is_header_row:
            ws_params.merge_cells(start_row=r_idx+1, start_column=1, end_row=r_idx+1, end_column=6)
            apply_style(ws_params.cell(row=r_idx+1, column=1), header_font, header_fill, Alignment(horizontal='center', vertical='center'))
    
    # Formattazione esplicita del foglio Parametri
    # Volumi giornalieri: B4:B7, D4:D7, E4:E7 -> interi
    for r in range(4, 8):
        ws_params.cell(row=r, column=2).number_format = '0'
        ws_params.cell(row=r, column=4).number_format = '0'
        ws_params.cell(row=r, column=5).number_format = '0'

    # Prezzi IVA Inclusa (B11:B13), IVA Esclusa (C11:C13) e costi unitari (D11:D13) -> valuta con due decimali
    for r in range(11, 14):
        ws_params.cell(row=r, column=2).number_format = '€ #,##0.00'  # IVA Inclusa
        ws_params.cell(row=r, column=3).number_format = '€ #,##0.00'  # IVA Esclusa
        ws_params.cell(row=r, column=4).number_format = '€ #,##0.00'  # Costo Unità

    # Margini percentuali: E11:E13 -> percentuale con due decimali
    for r in range(11, 14):
        ws_params.cell(row=r, column=5).number_format = '0.00%'

    # OPEX mensili: B17:B24 -> valuta senza decimali (aggiornato da 25 a 24)
    for r in range(17, 25):
        ws_params.cell(row=r, column=2).number_format = '€ #,##0'

    # Numero addetti e Giorni lavorativi: B19, B28 -> interi
    ws_params['B19'].number_format = '0'
    ws_params['B28'].number_format = '0'  # Aggiornato da B29 a B28
    
    # Set column widths
    ws_params.column_dimensions['A'].width = 35
    ws_params.column_dimensions['B'].width = 20
    ws_params.column_dimensions['C'].width = 20
    ws_params.column_dimensions['D'].width = 15
    ws_params.column_dimensions['E'].width = 15
    ws_params.column_dimensions['F'].width = 30

    # Blocca intestazione Parametri
    ws_params.freeze_panes = "A4"

    # Data validation e righe opzionali
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_volumi = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_prezzi = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_costi = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_giorni = DataValidation(type="whole", operator="between", formula1="1", formula2="31")
    dv_commissione = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    dv_altri_ricavi = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_sconti = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")

    ws_params.add_data_validation(dv_volumi)
    ws_params.add_data_validation(dv_prezzi)
    ws_params.add_data_validation(dv_costi)
    ws_params.add_data_validation(dv_giorni)
    ws_params.add_data_validation(dv_commissione)
    ws_params.add_data_validation(dv_altri_ricavi)
    ws_params.add_data_validation(dv_sconti)

    dv_volumi.add("B4:B7")
    dv_prezzi.add("C11:C13")  # IVA Esclusa (aggiornato da B11:B13)
    dv_costi.add("D11:D13")  # Costo Unità (aggiornato da C11:C13)
    dv_giorni.add("B28")  # Aggiornato da B29 a B28
    dv_commissione.add("B29")  # Aggiornato da B30 a B29
    
    # Aggiunta nuove righe parametri
    ws_params.cell(row=31, column=1, value="Altri ricavi mensili")
    ws_params.cell(row=31, column=2, value=0)
    ws_params.cell(row=31, column=3, value="€")
    ws_params.cell(row=32, column=1, value="Sconti su ricavi (%)")
    ws_params.cell(row=32, column=2, value=0.0)
    ws_params.cell(row=32, column=3, value="% sul ricavo")

    dv_altri_ricavi.add("B31")
    dv_sconti.add("B32")

    # Formati numerici
    ws_params['B29'].number_format = '0.00%'  # Aggiornato da B30 a B29
    ws_params['B31'].number_format = '€ #,##0'
    ws_params['B32'].number_format = '0.00%'

    # Sezione GARANZIE DI LOCAZIONE
    gar_title_row = 33
    ws_params.cell(row=gar_title_row, column=1, value="GARANZIE DI LOCAZIONE")
    ws_params.merge_cells(start_row=gar_title_row, start_column=1, end_row=gar_title_row, end_column=6)
    apply_style(ws_params.cell(row=gar_title_row, column=1), header_font, header_fill, Alignment(horizontal='center', vertical='center'))

    # Intestazioni della sezione garanzie
    ws_params.cell(row=gar_title_row + 1, column=1, value="Descrizione")
    ws_params.cell(row=gar_title_row + 1, column=2, value="Valore")
    ws_params.cell(row=gar_title_row + 1, column=3, value="Unità/Note")
    apply_style(ws_params.cell(row=gar_title_row + 1, column=1), subheader_font, subheader_fill)
    apply_style(ws_params.cell(row=gar_title_row + 1, column=2), subheader_font, subheader_fill)
    apply_style(ws_params.cell(row=gar_title_row + 1, column=3), subheader_font, subheader_fill)

    # Canone annuo di affitto = mensile * 12 (usa B17 come affitto mensile)
    ws_params.cell(row=gar_title_row + 2, column=1, value="Canone Annuo di Affitto")
    ws_params.cell(row=gar_title_row + 2, column=2, value="=Parametri!B17*12")
    ws_params.cell(row=gar_title_row + 2, column=3, value="€ annui")
    ws_params.cell(row=gar_title_row + 2, column=2).number_format = '€ #,##0'

    # Mensilità richieste a garanzia (valore modificabile)
    ws_params.cell(row=gar_title_row + 3, column=1, value="Mensilità richieste a Garanzia")
    ws_params.cell(row=gar_title_row + 3, column=2, value=12)
    ws_params.cell(row=gar_title_row + 3, column=3, value="mesi")
    ws_params.cell(row=gar_title_row + 3, column=2).number_format = '0'

    # Importo da Garantire = affitto mensile * mensilità
    ws_params.cell(row=gar_title_row + 4, column=1, value="Importo da Garantire")
    ws_params.cell(row=gar_title_row + 4, column=2, value=f"=Parametri!B17*B{gar_title_row + 3}")
    ws_params.cell(row=gar_title_row + 4, column=3, value="€")
    ws_params.cell(row=gar_title_row + 4, column=2).number_format = '€ #,##0'

    # Costo % annuo fideiussione (percentuale)
    ws_params.cell(row=gar_title_row + 5, column=1, value="Costo % Annuo Fideiussione")
    ws_params.cell(row=gar_title_row + 5, column=2, value=0.015)
    ws_params.cell(row=gar_title_row + 5, column=3, value="% annuo")
    ws_params.cell(row=gar_title_row + 5, column=2).number_format = '0.00%'

    # Validazione dati per la % annuale fideiussione (0-100%)
    dv_commissione.add(f"B{gar_title_row + 5}")

    # ==============================================================================
    # 2. FOGLIO P&L MENSILE (CON RAMP-UP 12 MESI)
    # ==============================================================================
    ws_pnl = wb.create_sheet("P&L Mensile")

    ws_pnl['A1'] = "PROFIT & LOSS MENSILE CON RAMP-UP (COLLEGATO AI PARAMETRI)"
    ws_pnl.merge_cells('A1:N1')
    apply_style(ws_pnl['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws_pnl.row_dimensions[1].height = 30

    # Intestazioni con 12 mesi
    headers_pnl = ["Voce"] + [f"Mese {i}" for i in range(1, 13)]
    for i, h in enumerate(headers_pnl, 1):
        cell = ws_pnl.cell(row=3, column=i, value=h)
        apply_style(cell, subheader_font, subheader_fill)
    # Blocca intestazione
    ws_pnl.freeze_panes = "A4"

    # Helper per applicare ramp-up coefficient
    def get_rampup_for_month(mese):
        """Restituisce il coefficiente di ramp-up per il mese (1-12)"""
        if mese <= 2:
            return 0.30
        elif mese <= 9:
            return 0.60
        else:
            return 1.00

    row_pnl = 4

    # Sezione RICAVI
    ws_pnl.cell(row=row_pnl, column=1, value="RICAVI").font = header_font
    row_pnl += 1

    # Piatti in store
    row_instore = row_pnl
    ws_pnl.cell(row=row_instore, column=1, value="Piatti pronti (in store)")
    for mese in range(1, 13):
        coeff = get_rampup_for_month(mese)
        ws_pnl.cell(row=row_instore, column=mese+1,
                    value=f"=Parametri!B4*Parametri!B28*Parametri!C11*{coeff}")
        ws_pnl.cell(row=row_instore, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Piatti delivery
    row_delivery = row_pnl
    ws_pnl.cell(row=row_delivery, column=1, value="Piatti pronti (delivery)")
    for mese in range(1, 13):
        coeff = get_rampup_for_month(mese)
        ws_pnl.cell(row=row_delivery, column=mese+1,
                    value=f"=Parametri!B5*Parametri!B28*Parametri!C11*{coeff}")
        ws_pnl.cell(row=row_delivery, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Bevande bibite
    row_bibite = row_pnl
    ws_pnl.cell(row=row_bibite, column=1, value="Bevande (bibite)")
    for mese in range(1, 13):
        coeff = get_rampup_for_month(mese)
        ws_pnl.cell(row=row_bibite, column=mese+1,
                    value=f"=Parametri!B6*Parametri!B28*Parametri!C12*{coeff}")
        ws_pnl.cell(row=row_bibite, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Bevande acqua
    row_acqua = row_pnl
    ws_pnl.cell(row=row_acqua, column=1, value="Bevande (acqua)")
    for mese in range(1, 13):
        coeff = get_rampup_for_month(mese)
        ws_pnl.cell(row=row_acqua, column=mese+1,
                    value=f"=Parametri!B7*Parametri!B28*Parametri!C13*{coeff}")
        ws_pnl.cell(row=row_acqua, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Altri ricavi (senza ramp-up, valore fisso)
    row_altri = row_pnl
    ws_pnl.cell(row=row_altri, column=1, value="Altri ricavi")
    for mese in range(1, 13):
        ws_pnl.cell(row=row_altri, column=mese+1, value="=Parametri!B31")
        ws_pnl.cell(row=row_altri, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Sconti su ricavi (proporzionali ai ricavi)
    row_sconti = row_pnl
    ws_pnl.cell(row=row_sconti, column=1, value="Sconti su ricavi")
    for mese in range(1, 13):
        col = get_column_letter(mese+1)
        ws_pnl.cell(row=row_sconti, column=mese+1,
                    value=f"=-SUM({col}{row_instore}:{col}{row_altri})*Parametri!B32")
        ws_pnl.cell(row=row_sconti, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Totale Ricavi
    row_ricavi_tot = row_pnl
    ws_pnl.cell(row=row_ricavi_tot, column=1, value="Ricavi Totali").font = total_font
    ws_pnl.cell(row=row_ricavi_tot, column=1).fill = total_fill
    for mese in range(1, 13):
        col = get_column_letter(mese+1)
        cell = ws_pnl.cell(row=row_ricavi_tot, column=mese+1,
                    value=f"=SUM({col}{row_instore}:{col}{row_sconti})")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_pnl += 1

    row_pnl += 1  # riga vuota

    # Sezione COGS
    ws_pnl.cell(row=row_pnl, column=1, value="COSTI VARIABILI (COGS)").font = header_font
    row_pnl += 1

    # Costo piatti pronti (con ramp-up)
    row_cogs_piatti = row_pnl
    ws_pnl.cell(row=row_cogs_piatti, column=1, value="Costo piatti pronti")
    for mese in range(1, 13):
        coeff = get_rampup_for_month(mese)
        ws_pnl.cell(row=row_cogs_piatti, column=mese+1,
                    value=f"=(Parametri!B4+Parametri!B5)*Parametri!B28*Parametri!D11*{coeff}")
        ws_pnl.cell(row=row_cogs_piatti, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Costo bibite (con ramp-up)
    row_cogs_bibite = row_pnl
    ws_pnl.cell(row=row_cogs_bibite, column=1, value="Costo bevande (bibite)")
    for mese in range(1, 13):
        coeff = get_rampup_for_month(mese)
        ws_pnl.cell(row=row_cogs_bibite, column=mese+1,
                    value=f"=Parametri!B6*Parametri!B28*Parametri!D12*{coeff}")
        ws_pnl.cell(row=row_cogs_bibite, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Costo acqua (con ramp-up)
    row_cogs_acqua = row_pnl
    ws_pnl.cell(row=row_cogs_acqua, column=1, value="Costo bevande (acqua)")
    for mese in range(1, 13):
        coeff = get_rampup_for_month(mese)
        ws_pnl.cell(row=row_cogs_acqua, column=mese+1,
                    value=f"=Parametri!B7*Parametri!B28*Parametri!D13*{coeff}")
        ws_pnl.cell(row=row_cogs_acqua, column=mese+1).number_format = '€ #,##0'
    row_pnl += 1

    # Totale COGS
    row_cogs_tot = row_pnl
    ws_pnl.cell(row=row_cogs_tot, column=1, value="Costi Variabili Totali (COGS)").font = total_font
    ws_pnl.cell(row=row_cogs_tot, column=1).fill = total_fill
    for mese in range(1, 13):
        col = get_column_letter(mese+1)
        cell = ws_pnl.cell(row=row_cogs_tot, column=mese+1,
                    value=f"=SUM({col}{row_cogs_piatti}:{col}{row_cogs_acqua})")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_pnl += 1

    row_pnl += 1  # riga vuota

    # Margine di contribuzione
    row_gross_margin = row_pnl
    ws_pnl.cell(row=row_gross_margin, column=1, value="Margine di Contribuzione (Gross Margin)").font = total_font
    ws_pnl.cell(row=row_gross_margin, column=1).fill = total_fill
    for mese in range(1, 13):
        col = get_column_letter(mese+1)
        cell = ws_pnl.cell(row=row_gross_margin, column=mese+1,
                    value=f"={col}{row_ricavi_tot}-{col}{row_cogs_tot}")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_pnl += 1

    row_pnl += 1  # riga vuota

    # Sezione OPEX
    ws_pnl.cell(row=row_pnl, column=1, value="COSTI OPERATIVI (OPEX)").font = header_font
    row_pnl += 1

    # OPEX fissi (identici ogni mese)
    opex_items = [
        ("Affitto locale", "=Parametri!B17"),
        ("Costo Personale", "=Parametri!B18"),
        ("Trasporti e logistica", "=Parametri!B21"),
        ("Utenze", "=Parametri!B22"),
        ("Marketing e pubblicità", "=Parametri!B23"),
    ]

    opex_rows = []
    for label, ref in opex_items:
        row_opex_item = row_pnl
        ws_pnl.cell(row=row_opex_item, column=1, value=label)
        for mese in range(1, 13):
            ws_pnl.cell(row=row_opex_item, column=mese+1, value=ref)
            ws_pnl.cell(row=row_opex_item, column=mese+1).number_format = '€ #,##0'
        opex_rows.append(row_opex_item)
        row_pnl += 1

    # Commissioni delivery (variano con le vendite delivery)
    row_comm_delivery = row_pnl
    ws_pnl.cell(row=row_comm_delivery, column=1, value="Commissioni delivery")
    for mese in range(1, 13):
        col = get_column_letter(mese+1)
        ws_pnl.cell(row=row_comm_delivery, column=mese+1,
                    value=f"={col}{row_delivery}*Parametri!B29")
        ws_pnl.cell(row=row_comm_delivery, column=mese+1).number_format = '€ #,##0'
    opex_rows.append(row_comm_delivery)
    row_pnl += 1

    # Costo Fideiussione Mensile
    row_fideius = row_pnl
    ws_pnl.cell(row=row_fideius, column=1, value="Costo Fideiussione Mensile")
    for mese in range(1, 13):
        ws_pnl.cell(row=row_fideius, column=mese+1, value="=(Parametri!B37*Parametri!B38)/12")
        ws_pnl.cell(row=row_fideius, column=mese+1).number_format = '€ #,##0'
    opex_rows.append(row_fideius)
    row_pnl += 1

    # Assicurazioni e varie
    row_assic = row_pnl
    ws_pnl.cell(row=row_assic, column=1, value="Assicurazioni e varie")
    for mese in range(1, 13):
        ws_pnl.cell(row=row_assic, column=mese+1, value="=Parametri!B25")
        ws_pnl.cell(row=row_assic, column=mese+1).number_format = '€ #,##0'
    opex_rows.append(row_assic)
    row_pnl += 1

    # Totale OPEX
    row_opex_tot = row_pnl
    ws_pnl.cell(row=row_opex_tot, column=1, value="Costi Operativi Totali (OPEX)").font = total_font
    ws_pnl.cell(row=row_opex_tot, column=1).fill = total_fill
    for mese in range(1, 13):
        col = get_column_letter(mese+1)
        cell = ws_pnl.cell(row=row_opex_tot, column=mese+1,
                    value=f"=SUM({col}{opex_rows[0]}:{col}{opex_rows[-1]})")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_pnl += 1

    row_pnl += 1  # riga vuota

    # EBITDA
    row_ebitda = row_pnl
    ws_pnl.cell(row=row_ebitda, column=1, value="EBITDA (Margine Operativo Lordo)").font = total_font
    ws_pnl.cell(row=row_ebitda, column=1).fill = total_fill
    for mese in range(1, 13):
        col = get_column_letter(mese+1)
        cell = ws_pnl.cell(row=row_ebitda, column=mese+1,
                    value=f"={col}{row_gross_margin}-{col}{row_opex_tot}")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font

    # Bandatura a zebra delle righe dati
    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Grigio chiaro
    zebra_index = 0
    for r in range(4, row_ebitda+1):
        label = ws_pnl.cell(row=r, column=1).value
        if label in (None, "RICAVI", "COSTI VARIABILI (COGS)", "COSTI OPERATIVI (OPEX)"):
            continue
        zebra_index += 1
        if zebra_index % 2 == 0:
            for c in range(1, 14):  # 1 label column + 12 month columns
                ws_pnl.cell(row=r, column=c).fill = light_fill

    # Larghezze colonne
    ws_pnl.column_dimensions['A'].width = 35
    for col in range(2, 14):  # Columns B to M (12 months)
        ws_pnl.column_dimensions[get_column_letter(col)].width = 13
    
    # ==============================================================================
    # 3. FOGLIO INVESTIMENTO & ROI
    # ==============================================================================
    ws_roi = wb.create_sheet("Investimento & ROI")

    ws_roi['A1'] = "ANALISI INVESTIMENTO INIZIALE E RITORNO"
    ws_roi.merge_cells('A1:D1')
    apply_style(ws_roi['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws_roi.row_dimensions[1].height = 30
    ws_roi.freeze_panes = "A3"
    
    investment_data = [
        ["COSTI DI STARTUP (CAPEX)", "Importo (€)", "% sul Totale", "Note"],
        ["Allestimento e ristrutturazione locale", 25000, "=B4/B12", "Opere murarie, impianti, design (25mq)"],
        ["Attrezzature cucina (6 microonde, frighi)", 8000, "=B5/B12", "Attrezzatura professionale"],
        ["Cassa automatica (16335+5360)", 21695, "=B6/B12", "Sistema POS e cassa automatica integrata"],
        ["Arredi, display e insegne", 10000, "=B7/B12", "Comunicazione in-store e esterna"],
        ["Costi burocratici e legali", 4000, "=B8/B12", "Licenze, consulenze, apertura società"],
        ["Budget Marketing di Lancio", 15000, "=B9/B12", "Campagna pre-opening e lancio"],
        ["Totale CAPEX", "=SUM(B4:B9)", "=B10/B12", ""],
        [],
        ["CAPITALE OPERATIVO INIZIALE"],
        ["Capitale Iniziale (3 Mesi OPEX)", f"='P&L Mensile'!M{row_opex_tot}*3", "=B11/B12", "Cassa per coprire i costi operativi dei primi 3 mesi"],
        ["INVESTIMENTO TOTALE RICHIESTO", "=B10+B11", "=B12/B12", "Fabbisogno finanziario per partire"],
        [],
        ["ANALISI DEL RITORNO SULL'INVESTIMENTO (ROI)"],
        ["EBITDA Mensile Medio (da P&L)", f"=AVERAGE('P&L Mensile'!B{row_ebitda}:M{row_ebitda})", "", "Margine operativo lordo medio mensile"],
        ["EBITDA Annuale", "=B15*12", "", "Proiezione a 12 mesi basata sulla media"],
        ["Payback Period (Mesi)", "=IFERROR(B12/B15,0)", "", "Mesi necessari per recuperare l'investimento"],
        ["ROI a 1 Anno", "=IFERROR(B16/B12,0)", "", "(EBITDA Annuale / Investimento Totale)"],
    ]
    
    for r_idx, row_data in enumerate(investment_data, 1):
        for c_idx, cell_data in enumerate(row_data, 1):
            cell = ws_roi.cell(row=r_idx + 2, column=c_idx, value=cell_data)
            if r_idx == 1: apply_style(cell, subheader_font, subheader_fill)
            if r_idx in [1, 9, 15]: cell.font = header_font  # Header rows (aggiornato)
            if r_idx in [8, 13]: cell.font = total_font  # Total rows (aggiornato)

            if c_idx == 2: cell.number_format = '€ #,##0'
            if c_idx == 3: cell.number_format = '0.0%'
            if r_idx == 18 and c_idx == 2: cell.number_format = '0.0 "mesi"'  # Payback (aggiornato)
            if r_idx == 19 and c_idx == 2: cell.number_format = '0.0%'  # ROI (aggiornato)

    ws_roi.column_dimensions['A'].width = 40
    ws_roi.column_dimensions['B'].width = 18
    ws_roi.column_dimensions['C'].width = 15
    ws_roi.column_dimensions['D'].width = 45

    # ==============================================================================
    # 4. FOGLIO ANALISI SCENARI
    # ==============================================================================
    ws_scenari = wb.create_sheet("Analisi Scenari")

    ws_scenari['A1'] = "ANALISI SCENARI - CONFRONTO BASE / PESSIMISTICO / OTTIMISTICO"
    ws_scenari.merge_cells('A1:G1')
    apply_style(ws_scenari['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws_scenari.row_dimensions[1].height = 30
    ws_scenari.freeze_panes = "A4"

    # Intestazioni
    scenari_headers = ["Metrica", "Scenario Base", "Scenario Pessimistico", "Scenario Ottimistico", "Delta Pess.", "Delta Ott.", "Sensitività"]
    for i, h in enumerate(scenari_headers, 1):
        cell = ws_scenari.cell(row=3, column=i, value=h)
        apply_style(cell, subheader_font, subheader_fill)

    row_sc = 4

    # SEZIONE VOLUMI
    ws_scenari.cell(row=row_sc, column=1, value="VOLUMI GIORNALIERI").font = header_font
    row_sc += 1

    # Piatti in store
    ws_scenari.cell(row=row_sc, column=1, value="Piatti in store (pz/gg)")
    ws_scenari.cell(row=row_sc, column=2, value="=Parametri!B4")
    ws_scenari.cell(row=row_sc, column=3, value="=Parametri!D4")
    ws_scenari.cell(row=row_sc, column=4, value="=Parametri!E4")
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/(D{row_sc}+C{row_sc})*2,0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '#,##0'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    # Piatti delivery
    ws_scenari.cell(row=row_sc, column=1, value="Piatti delivery (pz/gg)")
    ws_scenari.cell(row=row_sc, column=2, value="=Parametri!B5")
    ws_scenari.cell(row=row_sc, column=3, value="=Parametri!D5")
    ws_scenari.cell(row=row_sc, column=4, value="=Parametri!E5")
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/(D{row_sc}+C{row_sc})*2,0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '#,##0'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    row_sc += 1  # Riga vuota

    # SEZIONE PERFORMANCE ECONOMICA
    ws_scenari.cell(row=row_sc, column=1, value="PERFORMANCE ECONOMICA MENSILE").font = header_font
    row_sc += 1

    # Ricavi Totali - calcolati per ogni scenario
    row_ricavi_base = row_sc
    ws_scenari.cell(row=row_sc, column=1, value="Ricavi Totali")
    # Base
    ws_scenari.cell(row=row_sc, column=2, value=f"=(Parametri!B4*Parametri!B28*Parametri!C11)+(Parametri!B5*Parametri!B28*Parametri!C11)+(Parametri!B6*Parametri!B28*Parametri!C12)+(Parametri!B7*Parametri!B28*Parametri!C13)+Parametri!B31")  # Prezzo IVA Escl (B11→C11)
    # Pessimistico
    ws_scenari.cell(row=row_sc, column=3, value=f"=(Parametri!D4*Parametri!B28*Parametri!C11)+(Parametri!D5*Parametri!B28*Parametri!C11)+(Parametri!D6*Parametri!B28*Parametri!C12)+(Parametri!D7*Parametri!B28*Parametri!C13)+Parametri!B31")  # Prezzo IVA Escl (B11→C11)
    # Ottimistico
    ws_scenari.cell(row=row_sc, column=4, value=f"=(Parametri!E4*Parametri!B28*Parametri!C11)+(Parametri!E5*Parametri!B28*Parametri!C11)+(Parametri!E6*Parametri!B28*Parametri!C12)+(Parametri!E7*Parametri!B28*Parametri!C13)+Parametri!B31")  # Prezzo IVA Escl (B11→C11)
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '€ #,##0'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    # COGS
    row_cogs_sc = row_sc
    ws_scenari.cell(row=row_sc, column=1, value="COGS Totali")
    # Base
    ws_scenari.cell(row=row_sc, column=2, value=f"=((Parametri!B4+Parametri!B5)*Parametri!B28*Parametri!D11)+(Parametri!B6*Parametri!B28*Parametri!D12)+(Parametri!B7*Parametri!B28*Parametri!D13)")  # Costo Unità (C11→D11)
    # Pessimistico
    ws_scenari.cell(row=row_sc, column=3, value=f"=((Parametri!D4+Parametri!D5)*Parametri!B28*Parametri!D11)+(Parametri!D6*Parametri!B28*Parametri!D12)+(Parametri!D7*Parametri!B28*Parametri!D13)")  # Costo Unità (C11→D11)
    # Ottimistico
    ws_scenari.cell(row=row_sc, column=4, value=f"=((Parametri!E4+Parametri!E5)*Parametri!B28*Parametri!D11)+(Parametri!E6*Parametri!B28*Parametri!D12)+(Parametri!E7*Parametri!B28*Parametri!D13)")  # Costo Unità (C11→D11)
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '€ #,##0'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    # Gross Margin
    ws_scenari.cell(row=row_sc, column=1, value="Gross Margin")
    ws_scenari.cell(row=row_sc, column=2, value=f"=B{row_ricavi_base}-B{row_cogs_sc}")
    ws_scenari.cell(row=row_sc, column=3, value=f"=C{row_ricavi_base}-C{row_cogs_sc}")
    ws_scenari.cell(row=row_sc, column=4, value=f"=D{row_ricavi_base}-D{row_cogs_sc}")
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '€ #,##0'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_gross_sc = row_sc
    row_sc += 1

    # OPEX (costante per tutti gli scenari)
    row_opex_sc = row_sc
    ws_scenari.cell(row=row_sc, column=1, value="OPEX Totali")
    ws_scenari.cell(row=row_sc, column=2, value="=Parametri!B17+Parametri!B18+Parametri!B21+Parametri!B22+Parametri!B23+Parametri!B24+Parametri!B25+(Parametri!B37*Parametri!B38)/12")
    ws_scenari.cell(row=row_sc, column=3, value=f"=B{row_sc}")
    ws_scenari.cell(row=row_sc, column=4, value=f"=B{row_sc}")
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '€ #,##0'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    # EBITDA
    row_ebitda_sc = row_sc
    ws_scenari.cell(row=row_sc, column=1, value="EBITDA").font = total_font
    ws_scenari.cell(row=row_sc, column=2, value=f"=B{row_gross_sc}-B{row_opex_sc}")
    ws_scenari.cell(row=row_sc, column=3, value=f"=C{row_gross_sc}-C{row_opex_sc}")
    ws_scenari.cell(row=row_sc, column=4, value=f"=D{row_gross_sc}-D{row_opex_sc}")
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '€ #,##0'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    # Margine EBITDA %
    ws_scenari.cell(row=row_sc, column=1, value="Margine EBITDA %")
    ws_scenari.cell(row=row_sc, column=2, value=f"=IFERROR(B{row_ebitda_sc}/B{row_ricavi_base},0)")
    ws_scenari.cell(row=row_sc, column=3, value=f"=IFERROR(C{row_ebitda_sc}/C{row_ricavi_base},0)")
    ws_scenari.cell(row=row_sc, column=4, value=f"=IFERROR(D{row_ebitda_sc}/D{row_ricavi_base},0)")
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '0.0%'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    row_sc += 1  # Riga vuota

    # SEZIONE ROI
    ws_scenari.cell(row=row_sc, column=1, value="RITORNO SULL'INVESTIMENTO").font = header_font
    row_sc += 1

    # Payback Period
    ws_scenari.cell(row=row_sc, column=1, value="Payback Period (mesi)")
    ws_scenari.cell(row=row_sc, column=2, value=f"=IFERROR('Investimento & ROI'!B15/B{row_ebitda_sc},0)")  # Corretto B16→B15
    ws_scenari.cell(row=row_sc, column=3, value=f"=IFERROR('Investimento & ROI'!B15/C{row_ebitda_sc},0)")  # Corretto B16→B15
    ws_scenari.cell(row=row_sc, column=4, value=f"=IFERROR('Investimento & ROI'!B15/D{row_ebitda_sc},0)")  # Corretto B16→B15
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((C{row_sc}-D{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '0.0 "mesi"'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    # ROI Anno 1
    row_roi_sc = row_sc
    ws_scenari.cell(row=row_sc, column=1, value="ROI Anno 1 (%)")
    ws_scenari.cell(row=row_sc, column=2, value=f"=IFERROR((B{row_ebitda_sc}*12)/'Investimento & ROI'!B15,0)")  # Corretto B16→B15
    ws_scenari.cell(row=row_sc, column=3, value=f"=IFERROR((C{row_ebitda_sc}*12)/'Investimento & ROI'!B15,0)")  # Corretto B16→B15
    ws_scenari.cell(row=row_sc, column=4, value=f"=IFERROR((D{row_ebitda_sc}*12)/'Investimento & ROI'!B15,0)")  # Corretto B16→B15
    ws_scenari.cell(row=row_sc, column=5, value=f"=C{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=6, value=f"=D{row_sc}-B{row_sc}")
    ws_scenari.cell(row=row_sc, column=7, value=f"=IFERROR((D{row_sc}-C{row_sc})/B{row_sc},0)")
    for c in [2, 3, 4, 5, 6]: ws_scenari.cell(row=row_sc, column=c).number_format = '0.0%'
    ws_scenari.cell(row=row_sc, column=7).number_format = '0.0%'
    row_sc += 1

    # Spiegazioni Payback Period e ROI
    row_sc += 1
    ws_scenari.cell(row=row_sc, column=1, value="SPIEGAZIONE METRICHE DI INVESTIMENTO").font = header_font
    ws_scenari.merge_cells(start_row=row_sc, start_column=1, end_row=row_sc, end_column=7)
    apply_style(ws_scenari.cell(row=row_sc, column=1), header_font, header_fill, Alignment(horizontal='center', vertical='center'))
    row_sc += 1

    # Payback Period explanation
    ws_scenari.cell(row=row_sc, column=1, value="Payback Period (Periodo di Rientro):").font = Font(bold=True)
    ws_scenari.merge_cells(start_row=row_sc, start_column=1, end_row=row_sc, end_column=7)
    row_sc += 1

    payback_text = ("Il Payback Period indica quanto tempo (in mesi) è necessario per recuperare completamente "
                   "l'investimento iniziale attraverso i flussi di cassa generati dall'attività. "
                   "Si calcola dividendo l'investimento totale per l'EBITDA mensile. "
                   "Esempio: con un investimento di €166.695 e un EBITDA mensile di €51.530, "
                   "il payback period è di 3,2 mesi, quindi l'investimento si ripaga in circa 3 mesi.")

    ws_scenari.cell(row=row_sc, column=1, value=payback_text)
    ws_scenari.cell(row=row_sc, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws_scenari.merge_cells(start_row=row_sc, start_column=1, end_row=row_sc, end_column=7)
    ws_scenari.row_dimensions[row_sc].height = 60
    row_sc += 1
    row_sc += 1

    # ROI explanation
    ws_scenari.cell(row=row_sc, column=1, value="ROI Anno 1 (Return on Investment):").font = Font(bold=True)
    ws_scenari.merge_cells(start_row=row_sc, start_column=1, end_row=row_sc, end_column=7)
    row_sc += 1

    roi_text = ("Il ROI (Return on Investment) misura la redditività dell'investimento in termini percentuali. "
               "Si calcola dividendo l'EBITDA annuale (EBITDA mensile × 12) per l'investimento totale. "
               "Un ROI del 37,1% significa che il primo anno genererai un ritorno pari al 37,1% dell'investimento iniziale. "
               "Esempio: con €166.695 investiti e un EBITDA annuale di €618.360, il ROI è del 371%, "
               "indicando un ritorno eccellente sull'investimento.")

    ws_scenari.cell(row=row_sc, column=1, value=roi_text)
    ws_scenari.cell(row=row_sc, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws_scenari.merge_cells(start_row=row_sc, start_column=1, end_row=row_sc, end_column=7)
    ws_scenari.row_dimensions[row_sc].height = 60

    # Larghezze colonne (allargate per leggibilità)
    ws_scenari.column_dimensions['A'].width = 35
    ws_scenari.column_dimensions['B'].width = 22
    ws_scenari.column_dimensions['C'].width = 22
    ws_scenari.column_dimensions['D'].width = 22
    ws_scenari.column_dimensions['E'].width = 18
    ws_scenari.column_dimensions['F'].width = 18
    ws_scenari.column_dimensions['G'].width = 18

    # ==============================================================================
    # 5. FOGLIO CASH FLOW (PROIEZIONE 12 MESI)
    # ==============================================================================
    ws_cf = wb.create_sheet("Cash Flow 12 Mesi")

    ws_cf['A1'] = "PROIEZIONE CASH FLOW - 12 MESI"
    ws_cf.merge_cells('A1:N1')
    apply_style(ws_cf['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws_cf.row_dimensions[1].height = 30
    ws_cf.freeze_panes = "A4"

    # Intestazioni mesi
    cf_headers = ["Voce"] + [f"Mese {i}" for i in range(1, 13)]
    for i, h in enumerate(cf_headers, 1):
        cell = ws_cf.cell(row=3, column=i, value=h)
        apply_style(cell, subheader_font, subheader_fill)

    row_cf = 4

    # SEZIONE ENTRATE
    ws_cf.cell(row=row_cf, column=1, value="ENTRATE (CASH IN)").font = header_font
    ws_cf.cell(row=row_cf, column=1).fill = header_fill
    row_cf += 1

    # Ricavi da vendite
    row_ricavi_cf = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Ricavi da vendite")
    for mese in range(1, 13):
        # Riferimento alla colonna specifica del mese nel P&L Mensile
        col_pnl = get_column_letter(mese+1)
        ws_cf.cell(row=row_cf, column=mese+1, value=f"='P&L Mensile'!{col_pnl}{row_ricavi_tot}")
        ws_cf.cell(row=row_cf, column=mese+1).number_format = '€ #,##0'
    row_cf += 1

    # Totale Entrate
    row_tot_in = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Totale Entrate").font = total_font
    ws_cf.cell(row=row_cf, column=1).fill = positive_fill
    for mese in range(1, 13):
        col_cf = get_column_letter(mese+1)
        cell = ws_cf.cell(row=row_cf, column=mese+1, value=f"={col_cf}{row_ricavi_cf}")
        cell.number_format = '€ #,##0'
        cell.fill = positive_fill
        cell.font = positive_font
    row_cf += 1

    row_cf += 1  # Riga vuota

    # SEZIONE USCITE
    ws_cf.cell(row=row_cf, column=1, value="USCITE (CASH OUT)").font = header_font
    ws_cf.cell(row=row_cf, column=1).fill = header_fill
    row_cf += 1

    # COGS
    row_cogs_cf = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Costi variabili (COGS)")
    for mese in range(1, 13):
        col_pnl = get_column_letter(mese+1)
        ws_cf.cell(row=row_cf, column=mese+1, value=f"='P&L Mensile'!{col_pnl}{row_cogs_tot}")
        ws_cf.cell(row=row_cf, column=mese+1).number_format = '€ #,##0'
    row_cf += 1

    # OPEX
    row_opex_cf = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Costi operativi (OPEX)")
    for mese in range(1, 13):
        col_pnl = get_column_letter(mese+1)
        ws_cf.cell(row=row_cf, column=mese+1, value=f"='P&L Mensile'!{col_pnl}{row_opex_tot}")
        ws_cf.cell(row=row_cf, column=mese+1).number_format = '€ #,##0'
    row_cf += 1

    # Investimento iniziale (solo mese 1)
    row_invest_cf = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Investimento iniziale")
    ws_cf.cell(row=row_cf, column=2, value="='Investimento & ROI'!B15")
    ws_cf.cell(row=row_cf, column=2).number_format = '€ #,##0'
    for mese in range(2, 13):
        ws_cf.cell(row=row_cf, column=mese+1, value=0)
        ws_cf.cell(row=row_cf, column=mese+1).number_format = '€ #,##0'
    row_cf += 1

    # Totale Uscite
    row_tot_out = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Totale Uscite").font = total_font
    ws_cf.cell(row=row_cf, column=1).fill = negative_fill
    for mese in range(1, 13):
        cell = ws_cf.cell(row=row_cf, column=mese+1, value=f"={get_column_letter(mese+1)}{row_cogs_cf}+{get_column_letter(mese+1)}{row_opex_cf}+{get_column_letter(mese+1)}{row_invest_cf}")
        cell.number_format = '€ #,##0'
        cell.fill = negative_fill
        cell.font = negative_font
    row_cf += 1

    row_cf += 1  # Riga vuota

    # SEZIONE CASH FLOW NETTO
    ws_cf.cell(row=row_cf, column=1, value="CASH FLOW ANALYSIS").font = header_font
    ws_cf.cell(row=row_cf, column=1).fill = header_fill
    row_cf += 1

    # Cash Flow mensile
    row_cf_mensile = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Cash Flow Mensile").font = total_font
    ws_cf.cell(row=row_cf, column=1).fill = total_fill
    for mese in range(1, 13):
        cell = ws_cf.cell(row=row_cf, column=mese+1, value=f"={get_column_letter(mese+1)}{row_tot_in}-{get_column_letter(mese+1)}{row_tot_out}")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_cf += 1

    # Cash Flow cumulativo
    row_cf_cumul = row_cf
    ws_cf.cell(row=row_cf, column=1, value="Cash Flow Cumulativo").font = total_font
    ws_cf.cell(row=row_cf, column=1).fill = total_fill
    cell = ws_cf.cell(row=row_cf, column=2, value=f"=B{row_cf_mensile}")
    cell.number_format = '€ #,##0'
    cell.fill = total_fill
    cell.font = total_font
    for mese in range(2, 13):
        cell = ws_cf.cell(row=row_cf, column=mese+1, value=f"={get_column_letter(mese)}{row_cf_cumul}+{get_column_letter(mese+1)}{row_cf_mensile}")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_cf += 1

    # Break-even month (mese in cui CF cumulativo diventa positivo)
    row_cf += 1
    ws_cf.cell(row=row_cf, column=1, value="Note:")
    cell_note = ws_cf.cell(row=row_cf, column=2, value="Break-even quando CF Cumulativo > 0")
    cell_note.font = note_font
    cell_note.fill = note_fill
    row_cf += 1 # Spazio aggiuntivo prima del grafico

    # Larghezze colonne
    ws_cf.column_dimensions['A'].width = 30
    for col in range(2, 15):
        ws_cf.column_dimensions[get_column_letter(col)].width = 12

    # --- AGGIUNTA GRAFICO CASH FLOW CUMULATIVO ---
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import ChartLines

    chart = LineChart()
    chart.title = "Andamento Cash Flow Cumulativo (12 mesi)"
    chart.style = 13
    chart.height = 15
    chart.width = 25
    chart.legend = None

    # Dati (Y-axis)
    data = Reference(ws_cf, min_col=2, min_row=row_cf_cumul, max_col=13, max_row=row_cf_cumul)
    chart.add_data(data, titles_from_data=False)

    # Categorie (X-axis)
    cats = Reference(ws_cf, min_col=2, min_row=3, max_col=13, max_row=3)
    chart.set_categories(cats)

    # Stile asse Y (Valori)
    chart.y_axis.title = "Euro (€)"
    chart.y_axis.number_format = '€ #,##0'
    chart.y_axis.majorGridlines = ChartLines()

    # Stile asse X (Mesi)
    chart.x_axis.title = "Mese"
    chart.x_axis.tickLblPos = "low"

    ws_cf.add_chart(chart, "A" + str(row_cf + 1))

    # ==============================================================================
    # 5B. FOGLIO CASH FLOW 24 MESI MULTI-STORE (5 PUNTI VENDITA)
    # ==============================================================================
    ws_cf24 = wb.create_sheet("Cash Flow 24 Mesi Multi-Store")

    ws_cf24['A1'] = "CASH FLOW 24 MESI - ESPANSIONE MULTI-STORE (5 PUNTI VENDITA)"
    ws_cf24.merge_cells('A1:Z1')
    apply_style(ws_cf24['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws_cf24.row_dimensions[1].height = 30
    ws_cf24.freeze_panes = "A4"

    # Helper function per calcolare il coefficiente di ramp-up
    def get_rampup_coefficient(mese_corrente, mese_apertura):
        """
        Calcola il coefficiente di ramp-up in Python invece di usare formule IF annidate:
        - Mesi 1-2 dall'apertura: 30%
        - Mesi 3-9 dall'apertura: 60%
        - Mese 10+ dall'apertura: 100%
        """
        if mese_corrente < mese_apertura:
            return 0
        mesi_attivo = mese_corrente - mese_apertura + 1
        if mesi_attivo <= 2:
            return 0.30
        elif mesi_attivo <= 9:
            return 0.60
        else:
            return 1.0

    # Parametri aperture PdV
    row_cf24 = 3
    ws_cf24.cell(row=row_cf24, column=1, value="PARAMETRI APERTURE PUNTI VENDITA").font = header_font
    row_cf24 += 1

    aperture_pdv = [
        ("Punto Vendita 1", 1, "Apertura immediata (anno 1)"),
        ("Punto Vendita 2", 13, "Inizio anno 2"),
        ("Punto Vendita 3", 15, "2 mesi dopo Punto Vendita 2"),
        ("Punto Vendita 4", 17, "2 mesi dopo Punto Vendita 3"),
        ("Punto Vendita 5", 19, "2 mesi dopo Punto Vendita 4"),
    ]

    for pdv_name, mese_apertura, nota in aperture_pdv:
        ws_cf24.cell(row=row_cf24, column=1, value=pdv_name)
        ws_cf24.cell(row=row_cf24, column=2, value=f"Mese {mese_apertura}")
        ws_cf24.cell(row=row_cf24, column=3, value="='Investimento & ROI'!B15")  # Parametrico: riferimento a investimento totale
        ws_cf24.cell(row=row_cf24, column=3).number_format = '€ #,##0'
        ws_cf24.cell(row=row_cf24, column=4, value=nota)
        ws_cf24.cell(row=row_cf24, column=4).font = param_font
        row_cf24 += 1

    row_parametri_end = row_cf24
    row_cf24 += 1

    # Intestazioni mesi (24 colonne)
    ws_cf24.cell(row=row_cf24, column=1, value="CASH FLOW CONSOLIDATO").font = header_font
    row_cf24 += 1

    headers_24 = ["Voce"] + [f"M{i}" for i in range(1, 25)]
    for i, h in enumerate(headers_24, 1):
        cell = ws_cf24.cell(row=row_cf24, column=i, value=h)
        apply_style(cell, subheader_font, subheader_fill)

    row_headers = row_cf24
    row_cf24 += 1

    # === SEZIONE: TOTALI CONSOLIDATI ===
    row_start_consolidato = row_cf24

    # Investimenti totali per mese
    row_invest_tot = row_cf24
    ws_cf24.cell(row=row_invest_tot, column=1, value="Investimenti Totali").font = total_font
    ws_cf24.cell(row=row_invest_tot, column=1).fill = total_fill

    # Per ogni mese, somma gli investimenti di tutti i Punto Vendita che aprono in quel mese
    for mese in range(1, 25):
        formula_parts = []
        for idx, (_, mese_apertura, _) in enumerate(aperture_pdv, 1):
            if mese == mese_apertura:
                formula_parts.append("'Investimento & ROI'!B15")  # Corretto B16→B15

        if formula_parts:
            ws_cf24.cell(row=row_invest_tot, column=mese+1, value=f"={'+'.join(formula_parts)}")
        else:
            ws_cf24.cell(row=row_invest_tot, column=mese+1, value=0)
        ws_cf24.cell(row=row_invest_tot, column=mese+1).number_format = '€ #,##0'
    row_cf24 += 1

    # Ricavi totali (somma di tutti i Punto Vendita con ramp-up)
    row_ricavi_tot_24 = row_cf24
    ws_cf24.cell(row=row_ricavi_tot_24, column=1, value="Ricavi Totali").font = total_font
    ws_cf24.cell(row=row_ricavi_tot_24, column=1).fill = total_fill

    # Salvo le righe dei ricavi individuali dei Punto Vendita per sommarle dopo
    row_cf24 += 1

    # COGS totali
    row_cogs_tot_24 = row_cf24
    ws_cf24.cell(row=row_cogs_tot_24, column=1, value="COGS Totali").font = total_font
    ws_cf24.cell(row=row_cogs_tot_24, column=1).fill = total_fill
    row_cf24 += 1

    # OPEX totali
    row_opex_tot_24 = row_cf24
    ws_cf24.cell(row=row_opex_tot_24, column=1, value="OPEX Totali").font = total_font
    ws_cf24.cell(row=row_opex_tot_24, column=1).fill = total_fill
    row_cf24 += 1

    # Cash Flow Mensile
    row_cf_mensile_24 = row_cf24
    ws_cf24.cell(row=row_cf_mensile_24, column=1, value="Cash Flow Mensile").font = total_font
    ws_cf24.cell(row=row_cf_mensile_24, column=1).fill = total_fill
    for mese in range(1, 25):
        col = get_column_letter(mese+1)
        cell = ws_cf24.cell(row=row_cf_mensile_24, column=mese+1,
                     value=f"={col}{row_ricavi_tot_24}-{col}{row_invest_tot}-{col}{row_cogs_tot_24}-{col}{row_opex_tot_24}")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_cf24 += 1

    # Cash Flow Cumulativo
    row_cf_cumul_24 = row_cf24
    ws_cf24.cell(row=row_cf_cumul_24, column=1, value="Cash Flow Cumulativo").font = total_font
    ws_cf24.cell(row=row_cf_cumul_24, column=1).fill = total_fill
    cell = ws_cf24.cell(row=row_cf_cumul_24, column=2, value=f"=B{row_cf_mensile_24}")
    cell.number_format = '€ #,##0'
    cell.fill = total_fill
    cell.font = total_font

    for mese in range(2, 25):
        col_curr = get_column_letter(mese+1)
        col_prev = get_column_letter(mese)
        cell = ws_cf24.cell(row=row_cf_cumul_24, column=mese+1,
                     value=f"={col_prev}{row_cf_cumul_24}+{col_curr}{row_cf_mensile_24}")
        cell.number_format = '€ #,##0'
        cell.fill = total_fill
        cell.font = total_font
    row_cf24 += 2

    # === SEZIONE: DETTAGLIO PER CIASCUN Punto Vendita ===
    rows_ricavi_pdv = []  # Per poi sommare nel consolidato

    for idx, (pdv_name, mese_apertura, nota) in enumerate(aperture_pdv, 1):
        ws_cf24.cell(row=row_cf24, column=1, value=f"DETTAGLIO {pdv_name}").font = header_font
        row_cf24 += 1

        # Ricavi con ramp-up
        row_ricavi_pdv = row_cf24
        rows_ricavi_pdv.append(row_ricavi_pdv)
        ws_cf24.cell(row=row_ricavi_pdv, column=1, value=f"  Ricavi {pdv_name}")

        for mese in range(1, 25):
            # Calcola il coefficiente di ramp-up in Python
            coeff = get_rampup_coefficient(mese, mese_apertura)
            # Usa formula semplice moltiplicando per il coefficiente
            if coeff > 0:
                ws_cf24.cell(row=row_ricavi_pdv, column=mese+1, value=f"='P&L Mensile'!M{row_ricavi_tot}*{coeff}")
            else:
                ws_cf24.cell(row=row_ricavi_pdv, column=mese+1, value=0)
            ws_cf24.cell(row=row_ricavi_pdv, column=mese+1).number_format = '€ #,##0'
        row_cf24 += 1

        # COGS con ramp-up (proporzionale ai ricavi)
        row_cogs_pdv = row_cf24
        ws_cf24.cell(row=row_cogs_pdv, column=1, value=f"  COGS {pdv_name}")

        for mese in range(1, 25):
            coeff = get_rampup_coefficient(mese, mese_apertura)
            if coeff > 0:
                ws_cf24.cell(row=row_cogs_pdv, column=mese+1, value=f"='P&L Mensile'!M{row_cogs_tot}*{coeff}")
            else:
                ws_cf24.cell(row=row_cogs_pdv, column=mese+1, value=0)
            ws_cf24.cell(row=row_cogs_pdv, column=mese+1).number_format = '€ #,##0'
        row_cf24 += 1

        # OPEX fissi dal mese di apertura
        row_opex_pdv = row_cf24
        ws_cf24.cell(row=row_opex_pdv, column=1, value=f"  OPEX {pdv_name}")

        for mese in range(1, 25):
            # OPEX è fisso da quando il Punto Vendita apre
            if mese >= mese_apertura:
                ws_cf24.cell(row=row_opex_pdv, column=mese+1, value=f"='P&L Mensile'!M{row_opex_tot}")
            else:
                ws_cf24.cell(row=row_opex_pdv, column=mese+1, value=0)
            ws_cf24.cell(row=row_opex_pdv, column=mese+1).number_format = '€ #,##0'
        row_cf24 += 1

        # Cash Flow PdV
        row_cf_pdv = row_cf24
        cell = ws_cf24.cell(row=row_cf_pdv, column=1, value=f"  CF {pdv_name}")
        cell.font = Font(name='Arial', size=10, bold=True)  # Usa font più piccolo per i dettagli

        for mese in range(1, 25):
            col = get_column_letter(mese+1)
            invest = 0
            if mese == mese_apertura:
                invest = "'Investimento & ROI'!B15"  # Corretto B16→B15
            else:
                invest = "0"
            ws_cf24.cell(row=row_cf_pdv, column=mese+1,
                         value=f"={col}{row_ricavi_pdv}-{invest}-{col}{row_cogs_pdv}-{col}{row_opex_pdv}")
            ws_cf24.cell(row=row_cf_pdv, column=mese+1).number_format = '€ #,##0'
        row_cf24 += 2

    # Ora popolo i totali consolidati sommando i dettagli dei PdV
    # Ricavi Totali
    for mese in range(1, 25):
        col = get_column_letter(mese+1)
        sum_ricavi = "+".join([f"{col}{r}" for r in rows_ricavi_pdv])
        ws_cf24.cell(row=row_ricavi_tot_24, column=mese+1, value=f"={sum_ricavi}")
        ws_cf24.cell(row=row_ricavi_tot_24, column=mese+1).number_format = '€ #,##0'

    # COGS Totali (già calcolati sopra, sommiamo dalle righe COGS dei PdV)
    rows_cogs_pdv = [rows_ricavi_pdv[i] + 1 for i in range(5)]
    for mese in range(1, 25):
        col = get_column_letter(mese+1)
        sum_cogs = "+".join([f"{col}{r}" for r in rows_cogs_pdv])
        ws_cf24.cell(row=row_cogs_tot_24, column=mese+1, value=f"={sum_cogs}")
        ws_cf24.cell(row=row_cogs_tot_24, column=mese+1).number_format = '€ #,##0'

    # OPEX Totali
    rows_opex_pdv = [rows_ricavi_pdv[i] + 2 for i in range(5)]
    for mese in range(1, 25):
        col = get_column_letter(mese+1)
        sum_opex = "+".join([f"{col}{r}" for r in rows_opex_pdv])
        ws_cf24.cell(row=row_opex_tot_24, column=mese+1, value=f"={sum_opex}")
        ws_cf24.cell(row=row_opex_tot_24, column=mese+1).number_format = '€ #,##0'

    # === SEZIONE KPI MULTI-STORE ===
    ws_cf24.cell(row=row_cf24, column=1, value="KPI MULTI-STORE").font = header_font
    row_cf24 += 1

    # Numero Punto Vendita attivi per mese
    row_n_pdv = row_cf24
    ws_cf24.cell(row=row_n_pdv, column=1, value="N° Punto Vendita Attivi")
    for mese in range(1, 25):
        # Conta quanti Punto Vendita sono aperti in questo mese
        count_formula = "+".join([f"IF({mese}>={ma},1,0)" for _, ma, _ in aperture_pdv])
        ws_cf24.cell(row=row_n_pdv, column=mese+1, value=f"={count_formula}")
        ws_cf24.cell(row=row_n_pdv, column=mese+1).number_format = '0'
    row_cf24 += 1

    # Ricavi medi per Punto Vendita attivo
    row_ricavi_medi = row_cf24
    ws_cf24.cell(row=row_ricavi_medi, column=1, value="Ricavi Medi per Punto Vendita")
    for mese in range(1, 25):
        col = get_column_letter(mese+1)
        ws_cf24.cell(row=row_ricavi_medi, column=mese+1,
                     value=f"=IFERROR({col}{row_ricavi_tot_24}/{col}{row_n_pdv},0)")
        ws_cf24.cell(row=row_ricavi_medi, column=mese+1).number_format = '€ #,##0'
    row_cf24 += 1

    # EBITDA Margin % aggregato
    row_ebitda_margin = row_cf24
    ws_cf24.cell(row=row_ebitda_margin, column=1, value="EBITDA Margin %")
    for mese in range(1, 25):
        col = get_column_letter(mese+1)
        # EBITDA = Ricavi - COGS - OPEX
        ebitda = f"({col}{row_ricavi_tot_24}-{col}{row_cogs_tot_24}-{col}{row_opex_tot_24})"
        ws_cf24.cell(row=row_ebitda_margin, column=mese+1,
                     value=f"=IFERROR({ebitda}/{col}{row_ricavi_tot_24},0)")
        ws_cf24.cell(row=row_ebitda_margin, column=mese+1).number_format = '0.0%'
    row_cf24 += 1

    # Investimento totale cumulativo
    row_invest_cumul = row_cf24
    ws_cf24.cell(row=row_invest_cumul, column=1, value="Investimento Cumulativo")
    ws_cf24.cell(row=row_invest_cumul, column=2, value=f"=B{row_invest_tot}")
    ws_cf24.cell(row=row_invest_cumul, column=2).number_format = '€ #,##0'
    for mese in range(2, 25):
        col_curr = get_column_letter(mese+1)
        col_prev = get_column_letter(mese)
        ws_cf24.cell(row=row_invest_cumul, column=mese+1,
                     value=f"={col_prev}{row_invest_cumul}+{col_curr}{row_invest_tot}")
        ws_cf24.cell(row=row_invest_cumul, column=mese+1).number_format = '€ #,##0'
    row_cf24 += 2

    # === SEZIONE ANALISI FABBISOGNO FINANZIARIO ===
    row_cf24 += 1
    ws_cf24.cell(row=row_cf24, column=1, value="ANALISI FABBISOGNO FINANZIARIO").font = header_font
    row_cf24 += 1

    # Investimento totale teorico (5 Punto Vendita × investimento)
    ws_cf24.cell(row=row_cf24, column=1, value="Investimento Totale Teorico")
    ws_cf24.cell(row=row_cf24, column=2, value="=5*'Investimento & ROI'!B15")  # Corretto B16→B15
    ws_cf24.cell(row=row_cf24, column=2).number_format = '€ #,##0'
    ws_cf24.cell(row=row_cf24, column=3, value="=TEXT('Investimento & ROI'!B15,\"€#,##0\")&\" × 5 Punti Vendita\"")  # Parametrico
    ws_cf24.cell(row=row_cf24, column=3).font = param_font
    row_invest_teorico = row_cf24
    row_cf24 += 1

    # Punto di minimo CF (massimo fabbisogno) - CALCOLATO CON MIN()
    ws_cf24.cell(row=row_cf24, column=1, value="Punto di Minimo CF Cumulativo").font = total_font
    ws_cf24.cell(row=row_cf24, column=1).fill = total_fill
    # Formula MIN() sulla riga del Cash Flow Cumulativo (da colonna B a Y = 24 mesi)
    cell = ws_cf24.cell(row=row_cf24, column=2, value=f"=MIN(B{row_cf_cumul_24}:Y{row_cf_cumul_24})")
    cell.number_format = '€ #,##0'
    cell.fill = total_fill
    cell.font = total_font
    ws_cf24.cell(row=row_cf24, column=3, value="Massimo capitale richiesto (calcolato)")
    ws_cf24.cell(row=row_cf24, column=3).font = note_font
    ws_cf24.cell(row=row_cf24, column=3).fill = note_fill
    row_min_cf = row_cf24
    row_cf24 += 1

    # Investimento iniziale reale necessario (valore assoluto del minimo)
    ws_cf24.cell(row=row_cf24, column=1, value="INVESTIMENTO INIZIALE REALE").font = total_font
    ws_cf24.cell(row=row_cf24, column=1).fill = total_fill
    cell = ws_cf24.cell(row=row_cf24, column=2, value=f"=ABS(B{row_min_cf})")
    cell.number_format = '€ #,##0'
    cell.fill = positive_fill  # Verde per evidenziare il risultato positivo
    cell.font = positive_font
    ws_cf24.cell(row=row_cf24, column=3, value="Capitale da avere all'inizio")
    ws_cf24.cell(row=row_cf24, column=3).font = note_font
    ws_cf24.cell(row=row_cf24, column=3).fill = note_fill
    row_invest_reale = row_cf24
    row_cf24 += 1

    # Risparmio grazie al cash flow sequenziale
    ws_cf24.cell(row=row_cf24, column=1, value="Risparmio vs Investimento Teorico")
    ws_cf24.cell(row=row_cf24, column=2, value=f"=B{row_invest_teorico}-B{row_invest_reale}")
    ws_cf24.cell(row=row_cf24, column=2).number_format = '€ #,##0'
    ws_cf24.cell(row=row_cf24, column=3, value=f"=(B{row_invest_teorico}-B{row_invest_reale})/B{row_invest_teorico}")
    ws_cf24.cell(row=row_cf24, column=3).number_format = '0.0%'
    row_cf24 += 1

    # Mese in cui si raggiunge il minimo - USA MATCH() per trovarlo
    ws_cf24.cell(row=row_cf24, column=1, value="Mese del Massimo Fabbisogno")
    # MATCH trova la posizione del valore minimo nella riga CF Cumulativo
    ws_cf24.cell(row=row_cf24, column=2, value=f"=MATCH(B{row_min_cf},B{row_cf_cumul_24}:Y{row_cf_cumul_24},0)")
    ws_cf24.cell(row=row_cf24, column=2).number_format = '0'
    ws_cf24.cell(row=row_cf24, column=3, value="Mese con minimo CF cumulativo")
    ws_cf24.cell(row=row_cf24, column=3).font = param_font
    row_cf24 += 2

    # Note esplicative
    ws_cf24.cell(row=row_cf24, column=1, value="Note:")
    ws_cf24.cell(row=row_cf24, column=1).font = total_font
    row_cf24 += 1

    ws_cf24.cell(row=row_cf24, column=1, value="Ramp-up vendite:")
    ws_cf24.cell(row=row_cf24, column=2, value="30% (mesi 1-2), 60% (mesi 3-9), 100% (mese 10+)")
    ws_cf24.cell(row=row_cf24, column=2).font = param_font
    row_cf24 += 1

    ws_cf24.cell(row=row_cf24, column=1, value="Interpretazione:")
    ws_cf24.cell(row=row_cf24, column=2, value="L'investimento iniziale reale considera il cash flow positivo generato dal Punto Vendita 1")
    ws_cf24.cell(row=row_cf24, column=2).font = param_font
    row_cf24 += 1

    ws_cf24.cell(row=row_cf24, column=2, value="che copre parte dei costi di apertura dei Punto Vendita successivi.")
    ws_cf24.cell(row=row_cf24, column=2).font = param_font
    row_cf24 += 1

    ws_cf24.cell(row=row_cf24, column=2, value="NON servono €950,000 all'inizio, ma solo il capitale necessario per coprire")
    ws_cf24.cell(row=row_cf24, column=2).font = param_font
    row_cf24 += 1

    ws_cf24.cell(row=row_cf24, column=2, value="il punto di minimo del cash flow cumulativo.")
    ws_cf24.cell(row=row_cf24, column=2).font = param_font

    # Larghezze colonne
    ws_cf24.column_dimensions['A'].width = 35
    ws_cf24.column_dimensions['B'].width = 18
    ws_cf24.column_dimensions['C'].width = 18  # Ridotta da 35 a 18
    for col in range(4, 27):
        ws_cf24.column_dimensions[get_column_letter(col)].width = 11

    # ==============================================================================
    # 6. FOGLIO EXECUTIVE SUMMARY (IL RIEPILOGO FINALE)
    # ==============================================================================
    ws_summary = wb.create_sheet("Executive Summary", 0) # Messo come primo foglio
    
    ws_summary['A1'] = "EXECUTIVE SUMMARY - ITALIAN CORNER"
    ws_summary.merge_cells('A1:D1')
    apply_style(ws_summary['A1'], title_font, title_fill, Alignment(horizontal='center', vertical='center'))
    ws_summary.row_dimensions[1].height = 30
    ws_summary.freeze_panes = "A3"
    
    summary_data = [
        ["IL CONCEPT IN BREVE"],
        ["Modello di Business", "Catena di corner take-away per piatti pronti italiani di alta qualità, pronti in 60 secondi."],
        ["Unique Selling Proposition", "Tecnologia di conservazione a 12 mesi senza catena del freddo. Qualità ristorante, velocità fast-food."],
        ["Mercato Target", "Lavoratori (pranzo veloce), turisti (pasto tipico italiano), clienti salutisti (piatti bilanciati)."],
        [],
        ["METRICHE CHIAVE - SINGOLO PUNTO VENDITA (A REGIME)"],
        ["Investimento Iniziale per Punto Vendita", "='Investimento & ROI'!B16", "CAPEX completo per apertura."],
        ["Fatturato Mensile (100% capacità)", f"='P&L Mensile'!D{row_ricavi_tot}", "Dopo 10 mesi di ramp-up."],
        ["EBITDA Mensile (a regime)", f"='P&L Mensile'!D{row_ebitda}", "Margine operativo ricorrente."],
        ["Margine EBITDA %", f"='P&L Mensile'!E{row_ebitda}", "Redditività operativa."],
        ["Payback Period (singolo Punto Vendita)", "='Investimento & ROI'!B21", "Rientro investimento."],
        ["ROI Anno 1 (singolo Punto Vendita)", "='Investimento & ROI'!B22", "Ritorno primo anno."],
        [],
        ["PIANO DI ESPANSIONE 24 MESI (5 PUNTI VENDITA)"],
        ["Strategia", "Apertura sequenziale per autofinanziare l'espansione con cash flow dei primi PdV."],
        ["Timeline Aperture", "Punto Vendita 1 (Mese 1) → Punto Vendita 2 (Mese 13) → Punto Vendita 3-5 (ogni 2 mesi)."],
        ["Modello Ramp-Up", "30% capacità primi 2 mesi, 60% mesi 3-9, 100% da mese 10 in poi."],
        [],
        ["ANALISI FABBISOGNO FINANZIARIO (ESPANSIONE 5 Punti Vendita)"],
        ["Investimento Totale Teorico", f"='Cash Flow 24 Mesi Multi-Store'!B{row_invest_teorico}", "5 Punto Vendita × €190.000 (se aperti simultaneamente)."],
        ["Punto di Minimo CF Cumulativo", f"='Cash Flow 24 Mesi Multi-Store'!B{row_min_cf}", "Massimo scoperto necessario."],
        ["INVESTIMENTO INIZIALE REALE", f"='Cash Flow 24 Mesi Multi-Store'!B{row_invest_reale}", "Capitale effettivamente necessario all'inizio."],
        ["Risparmio vs Investimento Teorico", f"='Cash Flow 24 Mesi Multi-Store'!B{row_invest_reale + 1}", "Grazie al cash flow sequenziale."],
        ["% Risparmio", f"='Cash Flow 24 Mesi Multi-Store'!C{row_invest_reale + 1}", "Riduzione fabbisogno capitale."],
        [],
        ["RICHIESTA FINANZIARIA"],
        ["Importo Richiesto (SEED)", f"='Cash Flow 24 Mesi Multi-Store'!B{row_invest_reale}", "Basato sull'analisi cash flow 24 mesi."],
        ["Utilizzo dei Fondi", "Apertura sequenziale 5 Punto Vendita nei primi 24 mesi + capitale circolante."],
        ["Milestone di Validazione", "Break-even previsto entro Mese 20-22. Ricavi consolidati crescenti da Mese 13."],
    ]
    
    for r_idx, row_data in enumerate(summary_data, 1):
        if not row_data: continue
        for c_idx, cell_data in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r_idx + 2, column=c_idx, value=cell_data)
            if len(row_data) == 1:
                ws_summary.merge_cells(start_row=r_idx+2, start_column=1, end_row=r_idx+2, end_column=4)
                apply_style(cell, header_font, header_fill)
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if c_idx == 1:
                    cell.font = total_font
                    # Evidenzia in giallo "INVESTIMENTO INIZIALE REALE"
                    if "INVESTIMENTO INIZIALE REALE" in str(cell_data):
                        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

                if c_idx == 2 and isinstance(cell_data, str) and cell_data.startswith("="):
                    # Formattazione basata sul contenuto della formula
                    if ("E" in cell_data and str(row_ebitda) in cell_data) or ("B22" in cell_data) or ("!C" in cell_data and str(row_invest_reale + 1) in cell_data):
                        # Percentuali: Margine EBITDA, ROI, % Risparmio
                        cell.number_format = '0.0%'
                    elif "B21" in cell_data:
                        # Payback period
                        cell.number_format = '0.0 "mesi"'
                    else:
                        # Valori in euro
                        cell.number_format = '€ #,##0'

                    # Evidenzia in giallo l'importo richiesto
                    if f"!B{row_invest_reale}" in cell_data and "Importo Richiesto" in str(row_data[0]):
                        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

                if c_idx == 2 and isinstance(cell_data, (int, float)):
                    cell.number_format = '€ #,##0'
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 45
    ws_summary.column_dimensions['D'].width = 15

    # Grafici rimossi come da richiesta utente
    # Imposta il workbook per calcolare automaticamente
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # Marca tutti i fogli come "da ricalcolare"
    for sheet in wb.worksheets:
        sheet.sheet_properties.tabColor = None  # Reset properties

    # Salva il file
    file_path = "Business_Plan_Italian_Corner.xlsx"
    wb.save(file_path)

    # Chiudi e riapri per forzare il ricalcolo (workaround openpyxl)
    wb.close()

    return file_path

if __name__ == "__main__":
    file_generato = crea_business_plan_parametrizzato()
    print("\n✅ Complimenti, Paolo! Il tuo business plan dinamico è stato creato.")
    print(f"📍 Il file è stato salvato qui: {file_generato}")
    print("\n💡 PROSSIMI PASSI:")
    print("   1. Apri il file Excel.")
    print("   2. Vai al foglio 'Parametri' e modifica i valori nelle celle bianche per vedere come cambia l'intero modello.")
    print("   3. Il foglio 'Executive Summary' si aggiornerà automaticamente con i risultati finali.")