#!/usr/bin/env python3
"""Script per analizzare il file Excel del Business Plan"""

import openpyxl
from openpyxl.utils import get_column_letter

def analizza_business_plan():
    """Analizza il file Excel e stampa informazioni dettagliate"""

    file_path = "Business_Plan_Italian_Corner.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=False)

    print("="*80)
    print("ANALISI BUSINESS PLAN - ITALIAN CORNER")
    print("="*80)

    print(f"\n📊 Fogli presenti: {wb.sheetnames}")
    print(f"📄 Numero totale fogli: {len(wb.sheetnames)}\n")

    # Analizza ogni foglio
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print("\n" + "="*80)
        print(f"📋 FOGLIO: {sheet_name}")
        print("="*80)

        # Dimensioni del foglio
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"   Dimensioni: {max_row} righe x {max_col} colonne")

        # Celle con formule
        formule_count = 0
        valori_count = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formule_count += 1
                    else:
                        valori_count += 1

        print(f"   Celle con valori: {valori_count}")
        print(f"   Celle con formule: {formule_count}")

        # Mostra prime righe significative
        print(f"\n   📌 Prime righe significative:")
        rows_shown = 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, max_row), values_only=False), 1):
            row_values = []
            has_content = False
            for cell in row:
                if cell.value:
                    has_content = True
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        row_values.append(f"[FORMULA: {cell.value[:40]}...]" if len(cell.value) > 40 else f"[FORMULA: {cell.value}]")
                    else:
                        row_values.append(str(cell.value))
                else:
                    row_values.append("")

            if has_content and rows_shown < 15:
                print(f"      Riga {i}: {' | '.join(row_values[:4])}")  # Mostra solo prime 4 colonne
                rows_shown += 1

    # Analisi del foglio PARAMETRI
    print("\n" + "="*80)
    print("🔍 ANALISI DETTAGLIATA FOGLIO 'Parametri'")
    print("="*80)

    ws_params = wb["Parametri"]

    print("\n   📊 VOLUMI DI VENDITA (valori base):")
    print(f"      Piatti in store: {ws_params['B4'].value} pz/giorno")
    print(f"      Piatti delivery: {ws_params['B5'].value} pz/giorno")
    print(f"      Bevande bibite: {ws_params['B6'].value} pz/giorno")
    print(f"      Bevande acqua: {ws_params['B7'].value} pz/giorno")

    print("\n   💶 PREZZI DI VENDITA:")
    print(f"      Piatto pronto: €{ws_params['B11'].value}")
    print(f"      Bibita: €{ws_params['B12'].value}")
    print(f"      Acqua: €{ws_params['B13'].value}")

    print("\n   💰 COSTI UNITARI:")
    print(f"      Costo piatto: €{ws_params['C11'].value}")
    print(f"      Costo bibita: €{ws_params['C12'].value}")
    print(f"      Costo acqua: €{ws_params['C13'].value}")

    print("\n   🏢 OPEX MENSILI:")
    print(f"      Affitto: €{ws_params['B17'].value}")
    print(f"      Personale: {ws_params['B18'].value}")
    print(f"      N. addetti: {ws_params['B19'].value}")
    print(f"      Costo per addetto: €{ws_params['B20'].value}")
    print(f"      Trasporti: €{ws_params['B21'].value}")
    print(f"      Utenze: €{ws_params['B22'].value}")
    print(f"      Leasing: €{ws_params['B23'].value}")
    print(f"      Marketing: €{ws_params['B24'].value}")
    print(f"      Assicurazioni: €{ws_params['B25'].value}")

    print("\n   ⚙️ PARAMETRI GENERALI:")
    print(f"      Giorni lavorativi/mese: {ws_params['B29'].value}")
    print(f"      Commissione delivery: {ws_params['B30'].value}")

    # Analisi P&L
    print("\n" + "="*80)
    print("🔍 ANALISI DETTAGLIATA FOGLIO 'P&L Mensile'")
    print("="*80)

    ws_pnl = wb["P&L Mensile"]

    # Carica il workbook con data_only=True per leggere i valori calcolati
    wb_values = openpyxl.load_workbook(file_path, data_only=True)
    ws_pnl_values = wb_values["P&L Mensile"]

    print("\n   📈 METRICHE CHIAVE (valori calcolati):")

    # Cerca le righe con i totali (cerca per etichette)
    for row in range(1, ws_pnl.max_row + 1):
        cell_value = ws_pnl.cell(row=row, column=1).value
        if cell_value:
            if "Ricavi Totali" in str(cell_value):
                val = ws_pnl_values.cell(row=row, column=4).value
                pct = ws_pnl_values.cell(row=row, column=5).value
                print(f"      Ricavi Totali: €{val:,.0f}" if val else f"      Ricavi Totali: [FORMULA]")
            elif "Costi Variabili Totali" in str(cell_value) or "COGS" in str(cell_value):
                val = ws_pnl_values.cell(row=row, column=4).value
                pct = ws_pnl_values.cell(row=row, column=5).value
                if val:
                    print(f"      COGS: €{val:,.0f} ({pct:.1%})" if pct else f"      COGS: €{val:,.0f}")
                else:
                    print(f"      COGS: [FORMULA]")
            elif "Margine di Contribuzione" in str(cell_value) or "Gross Margin" in str(cell_value):
                val = ws_pnl_values.cell(row=row, column=4).value
                pct = ws_pnl_values.cell(row=row, column=5).value
                if val:
                    print(f"      Gross Margin: €{val:,.0f} ({pct:.1%})" if pct else f"      Gross Margin: €{val:,.0f}")
                else:
                    print(f"      Gross Margin: [FORMULA]")
            elif "Costi Operativi Totali" in str(cell_value) or "OPEX" in str(cell_value):
                val = ws_pnl_values.cell(row=row, column=4).value
                pct = ws_pnl_values.cell(row=row, column=5).value
                if val:
                    print(f"      OPEX Totali: €{val:,.0f} ({pct:.1%})" if pct else f"      OPEX Totali: €{val:,.0f}")
                else:
                    print(f"      OPEX Totali: [FORMULA]")
            elif "EBITDA" in str(cell_value):
                val = ws_pnl_values.cell(row=row, column=4).value
                pct = ws_pnl_values.cell(row=row, column=5).value
                if val:
                    print(f"      EBITDA: €{val:,.0f} ({pct:.1%})" if pct else f"      EBITDA: €{val:,.0f}")
                else:
                    print(f"      EBITDA: [FORMULA]")

    # Analisi Investimento & ROI
    print("\n" + "="*80)
    print("🔍 ANALISI DETTAGLIATA FOGLIO 'Investimento & ROI'")
    print("="*80)

    ws_roi = wb["Investimento & ROI"]
    ws_roi_values = wb_values["Investimento & ROI"]

    print("\n   💸 CAPEX E INVESTIMENTO:")
    for row in range(1, ws_roi.max_row + 1):
        cell_value = ws_roi.cell(row=row, column=1).value
        if cell_value:
            if "Totale CAPEX" in str(cell_value):
                val = ws_roi_values.cell(row=row, column=2).value
                print(f"      Totale CAPEX: €{val:,.0f}" if val else f"      Totale CAPEX: [FORMULA]")
            elif "INVESTIMENTO TOTALE" in str(cell_value):
                val = ws_roi_values.cell(row=row, column=2).value
                print(f"      Investimento Totale: €{val:,.0f}" if val else f"      Investimento Totale: [FORMULA]")
            elif "Payback Period" in str(cell_value):
                val = ws_roi_values.cell(row=row, column=2).value
                print(f"      Payback Period: {val:.1f} mesi" if val else f"      Payback Period: [FORMULA]")
            elif "ROI a 1 Anno" in str(cell_value):
                val = ws_roi_values.cell(row=row, column=2).value
                print(f"      ROI Anno 1: {val:.1%}" if val else f"      ROI Anno 1: [FORMULA]")

    print("\n" + "="*80)
    print("✅ ANALISI COMPLETATA")
    print("="*80)

if __name__ == "__main__":
    analizza_business_plan()
