#!/usr/bin/env python3
'''Script per analizzare il fabbisogno finanziario reale'''

import openpyxl

file_path = "Business_Plan_Italian_Corner.xlsx"

# Carica il workbook in due modalità:
# 1. data_only=False (default) per accedere alle formule e alla struttura
# 2. data_only=True per accedere ai valori calcolati (se disponibili)
try:
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    wb_values = openpyxl.load_workbook(file_path, data_only=True)
except FileNotFoundError:
    print(f"❌ Errore: Il file '{file_path}' non è stato trovato.")
    print("Assicurati di eseguire prima lo script 'crea_bp.py' per generarlo.")
    exit(1)


ws = wb_formulas["Cash Flow 24 Mesi Multi-Store"]
ws_val = wb_values["Cash Flow 24 Mesi Multi-Store"]


print("="*80)
print("ANALISI FABBISOGNO FINANZIARIO - ESPANSIONE 5 PdV")
print("="*80)

# Trova la riga del Cash Flow Cumulativo
row_cf_cumul = None
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val and "Cash Flow Cumulativo" in str(val):
        row_cf_cumul = r
        break

if not row_cf_cumul:
    print("❌ Riga Cash Flow Cumulativo non trovata!")
    exit(1)

print(f"\n📊 Riga Cash Flow Cumulativo: {row_cf_cumul}")

# Estrai i valori del CF cumulativo per tutti i 24 mesi
cf_values = []
for mese in range(1, 25):
    val = ws_val.cell(row=row_cf_cumul, column=mese+1).value
    cf_values.append((mese, val if val is not None else 0))

print("\n💰 CASH FLOW CUMULATIVO PER MESE:")
print("-" * 80)

# Mostra i valori in formato tabellare (4 mesi per riga)
for i in range(0, 24, 4):
    row_str = []
    for j in range(4):
        if i + j < 24:
            mese, valore = cf_values[i + j]
            row_str.append(f"M{mese:2d}: €{valore:>12,.0f}")
    print("  ".join(row_str))

# Trova il punto di minimo (massimo fabbisogno)
min_cf = min(cf_values, key=lambda x: x[1])
max_cf = max(cf_values, key=lambda x: x[1])

print("\n" + "="*80)
print("📉 ANALISI DEL FABBISOGNO")
print("="*80)

print(f"\n🔻 PUNTO DI MINIMO (Massimo Fabbisogno):")
print(f"   Mese: {min_cf[0]}")
print(f"   Cash Flow Cumulativo: €{min_cf[1]:,.0f}")
print(f"   👉 INVESTIMENTO INIZIALE NECESSARIO: €{abs(min_cf[1]):,.0f}")

print(f"\n🔺 PUNTO DI MASSIMO (Fine periodo):")
print(f"   Mese: {max_cf[0]}")
print(f"   Cash Flow Cumulativo: €{max_cf[1]:,.0f}")

# Trova il break-even (quando CF diventa positivo)
breakeven_mese = None
for mese, val in cf_values:
    if val > 0:
        breakeven_mese = mese
        break

if breakeven_mese:
    print(f"\n✅ BREAK-EVEN:")
    print(f"   Mese: {breakeven_mese}")
    print(f"   Il cash flow diventa positivo al mese {breakeven_mese}")
else:
    print(f"\n❌ BREAK-EVEN:")
    print(f"   Non raggiunto nei 24 mesi")

# Analisi per trimestre
print("\n" + "="*80)
print("📊 ANALISI PER TRIMESTRE")
print("="*80)

trimestri = [
    ("Q1 (M1-M3)", cf_values[2][1]),
    ("Q2 (M4-M6)", cf_values[5][1]),
    ("Q3 (M7-M9)", cf_values[8][1]),
    ("Q4 (M10-M12)", cf_values[11][1]),
    ("Q5 (M13-M15)", cf_values[14][1]),
    ("Q6 (M16-M18)", cf_values[17][1]),
    ("Q7 (M19-M21)", cf_values[20][1]),
    ("Q8 (M22-M24)", cf_values[23][1]),
]

for label, val in trimestri:
    print(f"   {label}: €{val:>12,.0f}")

# Trova gli investimenti
print("\n" + "="*80)
print("💸 INVESTIMENTI PIANIFICATI")
print("="*80)

row_invest = None
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val and "Investimenti Totali" in str(val):
        row_invest = r
        break

total_invest = 0
if row_invest:
    print("\nMese | Investimento")
    print("-" * 30)
    for mese in range(1, 25):
        val = ws_val.cell(row=row_invest, column=mese+1).value
        if val and val > 0:
            print(f"M{mese:2d}  | €{val:>12,.0f}")
            total_invest += val
    print("-" * 30)
    print(f"TOT  | €{total_invest:>12,.0f}")

print("\n" + "="*80)
print("💡 CONCLUSIONI")
print("="*80)

if total_invest > 0:
    print(f'''
Per aprire 5 punti vendita nei prossimi 24 mesi, NON hai bisogno di
€{total_invest:,.0f} (5 × ~€{total_invest/5:,.0f}) fin dall\'inizio!

Il cash flow generato dal PdV 1 (e successivamente dagli altri) copre
parte dei costi dei PdV successivi.

🎯 INVESTIMENTO INIZIALE REALE NECESSARIO: €{abs(min_cf[1]):,.0f}

Questo è il capitale che devi avere disponibile all\'inizio per coprire
il punto di massimo fabbisogno (mese {min_cf[0]}).

Differenza vs investimento totale: €{total_invest - abs(min_cf[1]):,.0f}
Risparmio: {((total_invest - abs(min_cf[1])) / total_invest * 100):.1f}%
''')
else:
    print('''
I dati sugli investimenti non sono stati trovati o sono pari a zero.
L\'analisi del risparmio non può essere completata.
Verifica che il file Excel contenga i dati corretti.
''')


print("="*80)
