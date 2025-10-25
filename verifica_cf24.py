#!/usr/bin/env python3
"""Script per verificare il foglio Cash Flow 24 Mesi Multi-Store"""

import openpyxl

file_path = "Business_Plan_Italian_Corner.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
wb_values = openpyxl.load_workbook(file_path, data_only=True)

ws = wb["Cash Flow 24 Mesi Multi-Store"]
ws_val = wb_values["Cash Flow 24 Mesi Multi-Store"]

print("="*80)
print("VERIFICA FOGLIO 'Cash Flow 24 Mesi Multi-Store'")
print("="*80)

print(f"\n📊 Dimensioni: {ws.max_row} righe x {ws.max_column} colonne")

# Conta formule e valori
formule = 0
valori = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formule += 1
            else:
                valori += 1

print(f"📈 Celle con valori: {valori}")
print(f"🔢 Celle con formule: {formule}")

# Mostra parametri aperture
print("\n" + "="*80)
print("PARAMETRI APERTURE PUNTI VENDITA")
print("="*80)

for r in range(4, 9):
    pdv = ws.cell(row=r, column=1).value
    mese = ws.cell(row=r, column=2).value
    invest = ws.cell(row=r, column=3).value
    nota = ws.cell(row=r, column=4).value
    if invest:
        print(f"{pdv}: {mese} - Investimento: {invest} - {nota}")
    else:
        print(f"{pdv}: {mese} - {nota}")

# Trova e mostra il consolidato
print("\n" + "="*80)
print("CASH FLOW CONSOLIDATO (campione primi 6 mesi)")
print("="*80)

# Cerca la riga del consolidato
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val and "CASH FLOW CONSOLIDATO" in str(val):
        row_consolidato_start = r + 2
        print(f"\nRiga inizio dati consolidato: {row_consolidato_start}")

        # Mostra le prime righe del consolidato per i primi 6 mesi
        for offset in range(7):  # Invest, Ricavi, COGS, OPEX, CF Mens, CF Cumul + header
            row_idx = row_consolidato_start + offset
            voce = ws.cell(row=row_idx, column=1).value
            if voce:
                print(f"\n{voce}:")
                values = []
                for m in range(1, 7):  # Primi 6 mesi
                    cell_val = ws_val.cell(row=row_idx, column=m+1).value
                    if cell_val is not None:
                        if isinstance(cell_val, (int, float)):
                            values.append(f"M{m}: €{cell_val:,.0f}")
                        else:
                            values.append(f"M{m}: {cell_val}")
                    else:
                        # Mostra la formula
                        formula = ws.cell(row=row_idx, column=m+1).value
                        if formula and isinstance(formula, str) and formula.startswith('='):
                            values.append(f"M{m}: [FORMULA]")
                        else:
                            values.append(f"M{m}: -")

                print("  " + " | ".join(values[:3]))
                if len(values) > 3:
                    print("  " + " | ".join(values[3:]))
        break

# Mostra dettaglio PdV
print("\n" + "="*80)
print("DETTAGLIO PUNTI VENDITA (campione Mesi 1, 12, 14)")
print("="*80)

for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val and "DETTAGLIO PdV" in str(val):
        pdv_num = val.split()[-1]
        print(f"\n{val}:")

        # Ricavi
        ricavi_row = r + 1
        ricavi_m1 = ws_val.cell(row=ricavi_row, column=2).value
        ricavi_m12 = ws_val.cell(row=ricavi_row, column=13).value
        ricavi_m14 = ws_val.cell(row=ricavi_row, column=15).value

        print(f"  Ricavi: M1=€{ricavi_m1:,.0f if ricavi_m1 else 0}, M12=€{ricavi_m12:,.0f if ricavi_m12 else 0}, M14=€{ricavi_m14:,.0f if ricavi_m14 else 0}")

        # CF
        cf_row = r + 4
        cf_m1 = ws_val.cell(row=cf_row, column=2).value
        cf_m12 = ws_val.cell(row=cf_row, column=13).value
        cf_m14 = ws_val.cell(row=cf_row, column=15).value

        print(f"  Cash Flow: M1=€{cf_m1:,.0f if cf_m1 else 0}, M12=€{cf_m12:,.0f if cf_m12 else 0}, M14=€{cf_m14:,.0f if cf_m14 else 0}")

# Mostra KPI
print("\n" + "="*80)
print("KPI MULTI-STORE (campione primi 6 mesi + mesi 12, 18, 24)")
print("="*80)

for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val and "KPI MULTI-STORE" in str(val):
        kpi_start = r + 1

        kpi_labels = ["N° PdV Attivi", "Ricavi Medi per PdV", "EBITDA Margin %", "Investimento Cumulativo"]

        for i, label in enumerate(kpi_labels):
            row_idx = kpi_start + i
            label_cell = ws.cell(row=row_idx, column=1).value
            if label_cell:
                vals = []
                for m in [1, 2, 3, 12, 18, 24]:
                    v = ws_val.cell(row=row_idx, column=m+1).value
                    if v is not None:
                        if "%" in str(label_cell):
                            vals.append(f"M{m}: {v:.1%}" if isinstance(v, float) else f"M{m}: {v}")
                        elif "PdV" in str(label_cell) and "N°" in str(label_cell):
                            vals.append(f"M{m}: {int(v)}" if isinstance(v, (int, float)) else f"M{m}: {v}")
                        else:
                            vals.append(f"M{m}: €{v:,.0f}" if isinstance(v, (int, float)) else f"M{m}: {v}")
                    else:
                        vals.append(f"M{m}: -")

                print(f"\n{label_cell}:")
                print("  " + " | ".join(vals))
        break

print("\n" + "="*80)
print("✅ VERIFICA COMPLETATA")
print("="*80)
